from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_step40_custom_plan_")

from app.db import init_db
from app.services import accounting, reporting, repository as repo
from app.services.parser import parse_message


def _save(chat_id: int, user_id: int, text: str) -> None:
    ops, errors = parse_message(chat_id, chat_id, text)
    assert not errors, (text, errors)
    saved = accounting.apply_operations(chat_id, chat_id, user_id, [op.to_dict() for op in ops], text)
    assert saved == 1, (text, saved, [op.to_dict() for op in ops])


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def main() -> None:
    init_db()
    chat_id = -4001
    user_id = 4001
    repo.set_chat_connected(chat_id, "Группа", "supergroup", True)

    for name in ("Комплектующая 1", "Комплектующая 2", "Комплектующая 3"):
        ok, msg = repo.create_entity(chat_id, "component", name, "шт")
        assert ok, msg
    for name in ("Изделие 1", "Изделие 2"):
        ok, msg = repo.create_entity(chat_id, "product", name, "шт")
        assert ok, msg

    products = {item.name: item for item in repo.list_entities(chat_id, {"product"})}
    comps = {item.name: item for item in repo.list_entities(chat_id, {"component"})}
    repo.set_product_components(chat_id, products["Изделие 1"].id, [(comps["Комплектующая 1"].id, 2), (comps["Комплектующая 2"].id, 1)])
    repo.set_product_components(chat_id, products["Изделие 2"].id, [(comps["Комплектующая 1"].id, 1), (comps["Комплектующая 3"].id, 3)])

    _save(chat_id, user_id, "Производство Комплектующая 1 1000000")
    _save(chat_id, user_id, "Производство Комплектующая 2 100000")
    _save(chat_id, user_id, "Производство Комплектующая 3 20000")
    _save(chat_id, user_id, "Собрали Изделие 1 10000")
    _save(chat_id, user_id, "На отправку Изделие 1 7000")

    request = "план сборки: Изделие 1 2500000; Изделие 2 50000"
    text = reporting.build_assembly_capacity_report(chat_id, request)
    assert "Изделие 1" in text and "2 500 000" in text, text
    assert "Изделие 2" in text and "50 000" in text, text
    assert "4 020 000" in text, text
    assert "4e+06" not in text and "e+" not in text.lower(), text

    xlsx = reporting.create_xlsx_report(chat_id, "excel отчёт за сегодня для сборки: Изделие 1 2500000; Изделие 2 50000", user_id=user_id)
    wb = load_workbook(xlsx, data_only=True, read_only=False)
    assert "План сборки" in wb.sheetnames, wb.sheetnames
    plan = wb["План сборки"]
    rows = _rows(plan)
    assert any(row[0] == "Изделие 1" and row[1] == 2500000 and row[3] == "Комплектующая 1" and row[7] == 4020000 for row in rows), rows
    assert any(row[0] == "Изделие 2" and row[1] == 50000 and row[3] == "Комплектующая 3" and row[7] == 130000 for row in rows), rows
    assert not any(row[0] == "Изделие 1" and row[1] == 10000 for row in rows), rows
    assert not any(row[0] == "Изделие 2" and row[1] == 100000 for row in rows), rows
    for cell in plan["H"]:
        if cell.row > 1 and isinstance(cell.value, (int, float)):
            assert "E" not in str(cell.value).upper(), cell.value
            assert cell.number_format != "General", cell.number_format

    txt = reporting.create_txt_report(chat_id, "txt отчёт за сегодня для сборки: Изделие 1 2500000; Изделие 2 50000", user_id=user_id)
    txt_data = txt.read_text("utf-8")
    assert "4 020 000" in txt_data, txt_data
    assert "4e+06" not in txt_data and "e+" not in txt_data.lower(), txt_data

    html = reporting.create_html_report(chat_id, "html отчёт за сегодня для сборки: Изделие 1 2500000, Изделие 2 50000", user_id=user_id)
    html_data = html.read_text("utf-8")
    assert "4 020 000" in html_data, html_data
    assert "4e+06" not in html_data and "e+" not in html_data.lower(), html_data
    print("OK")


if __name__ == "__main__":
    main()
