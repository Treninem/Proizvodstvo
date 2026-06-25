from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_multi_report_")

from app.db import init_db
from app.services import accounting, reporting, repository as repo
from app.services.parser import parse_message


def seed(chat_id: int, component_name: str, qty: int) -> None:
    repo.set_chat_connected(chat_id, f"Группа {abs(chat_id)}", "supergroup", True)
    ok, _ = repo.create_entity(chat_id, "component", component_name, "шт")
    assert ok
    ops, errors = parse_message(chat_id, chat_id, f"Производство {component_name} {qty}")
    assert not errors, errors
    accounting.apply_operations(chat_id, chat_id, 111, [op.to_dict() for op in ops], f"Производство {component_name} {qty}")


def main() -> None:
    init_db()
    seed(-3101, "Деталь А", 1000)
    seed(-3102, "Деталь Б", 2000)
    text = reporting.build_multi_text_report((-3101, -3102), "отчёт за сегодня", titles={-3101: "Цех 1", -3102: "Цех 2"}, user_id=111)
    assert "Общий отчёт" in text, text
    assert "Цех 1" in text and "Цех 2" in text, text
    assert "Деталь А" in text and "Деталь Б" in text, text
    xlsx = reporting.create_multi_xlsx_report((-3101, -3102), "отчёт за сегодня", titles={-3101: "Цех 1", -3102: "Цех 2"}, user_id=111)
    wb = load_workbook(xlsx, read_only=True)
    assert "Общий отчёт" in wb.sheetnames[0], wb.sheetnames
    assert len(wb.sheetnames) >= 3, wb.sheetnames
    assert any("Цех 1" in name for name in wb.sheetnames), wb.sheetnames
    assert any("Цех 2" in name for name in wb.sheetnames), wb.sheetnames
    print("OK")


if __name__ == "__main__":
    main()
