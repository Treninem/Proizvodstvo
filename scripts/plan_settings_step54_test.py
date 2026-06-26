from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_step54_plan_")

from app.db import init_db
from app.services import accounting, reporting, repository as repo
from app.services.parser import parse_message


def _save(chat_id: int, user_id: int, text: str) -> None:
    ops, errors = parse_message(chat_id, chat_id, text)
    assert not errors, (text, errors)
    saved = accounting.apply_operations(chat_id, chat_id, user_id, [op.to_dict() for op in ops], text)
    assert saved == 1, (text, saved)


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def main() -> None:
    init_db()
    chat_id = -1003940000000
    user_id = 7306158154
    repo.set_chat_connected(chat_id, "Цех готовой продукции", "supergroup", True)

    for name in ("Деталь А", "Деталь Б"):
        ok, msg = repo.create_entity(chat_id, "component", name, "шт")
        assert ok, msg
    ok, msg = repo.create_entity(chat_id, "product", "Изделие 1", "шт")
    assert ok, msg
    product = repo.list_entities(chat_id, {"product"})[0]
    comps = repo.list_entities(chat_id, {"component"})
    repo.set_product_components(chat_id, product.id, [(comps[0].id, 2), (comps[1].id, 1)])

    _save(chat_id, user_id, "Производство Деталь А 20000")
    _save(chat_id, user_id, "Производство Деталь Б 5000")

    # Без сохранённого плана встроенные цели не должны появляться.
    xlsx = reporting.create_xlsx_report(chat_id, "excel отчёт за сегодня", user_id=user_id)
    wb = load_workbook(xlsx, data_only=True, read_only=False)
    plan_rows = _rows(wb["План сборки"])
    assert not any(row[0] == "Изделие 1" and row[1] in (10000, 50000, 100000) for row in plan_rows), plan_rows

    repo.set_assembly_plan_targets(chat_id, product.id, [4000, 7000])
    text = reporting.build_assembly_capacity_report(chat_id, "план сборки")
    assert "7 000" in text and "не хватает 2 000" in text, text

    xlsx2 = reporting.create_xlsx_report(chat_id, "excel отчёт за сегодня", user_id=user_id)
    wb2 = load_workbook(xlsx2, data_only=True, read_only=False)
    journal = _rows(wb2["Журнал"])
    assert any(row[-1] == "Цех готовой продукции" for row in journal[1:]), journal
    assert not any("E+" in str(row[-1]).upper() for row in journal[1:]), journal
    plan = _rows(wb2["План сборки"])
    assert any(row[0] == "Изделие 1" and row[1] == 4000 for row in plan), plan
    assert any(row[0] == "Изделие 1" and row[1] == 7000 for row in plan), plan

    msgs = reporting.completed_plan_messages(chat_id)
    assert len(msgs) == 1, msgs
    assert "Изделие 1" in msgs[0] and "4 000" in msgs[0], msgs
    assert reporting.completed_plan_messages(chat_id) == []
    print("OK")


if __name__ == "__main__":
    main()
