from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_step38_reports_")

from app.db import init_db
from app.services import accounting, reporting, repository as repo
from app.services.parser import parse_message


def _save(chat_id: int, user_id: int, text: str) -> None:
    ops, errors = parse_message(chat_id, chat_id, text)
    assert not errors, (text, errors)
    saved = accounting.apply_operations(chat_id, chat_id, user_id, [op.to_dict() for op in ops], text)
    assert saved == 1, (text, saved, [op.to_dict() for op in ops])


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def main() -> None:
    init_db()
    chat_id = -3801
    user_id = 3801
    repo.set_chat_connected(chat_id, "Группа", "supergroup", True)

    for name in ("Часть А", "Часть Б"):
        ok, _ = repo.create_entity(chat_id, "component", name, "шт")
        assert ok
    ok, _ = repo.create_entity(chat_id, "product", "Изделие 1", "шт")
    assert ok
    product = repo.list_entities(chat_id, {"product"})[0]
    comps = repo.list_entities(chat_id, {"component"})
    repo.set_product_components(chat_id, product.id, [(comps[0].id, 2), (comps[1].id, 1)])

    _save(chat_id, user_id, "Сделали Часть А 30000")
    _save(chat_id, user_id, "Сделали Часть Б 16000")
    _save(chat_id, user_id, "Собрали Изделие 1 5000")
    _save(chat_id, user_id, "Зафасовали Изделие 1 2000")

    text = reporting.build_assembly_capacity_report(chat_id, "сколько нужно для сборки 50000 Изделие 1")
    assert "Итого сейчас: 10000 шт." in text, text
    assert "50000" in text and "не хватает 80000" in text and "не хватает 39000" in text, text

    xlsx = reporting.create_xlsx_report(chat_id, "excel отчёт за сегодня для сборки 250000", user_id=user_id)
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    required = {
        "Отчёт",
        "Склад",
        "Остатки комплектующих",
        "Комплектующие по датам",
        "Сборка и отправка по датам",
        "Расчёт сборки",
        "План сборки",
        "Собрано и отправлено",
    }
    assert required.issubset(set(wb.sheetnames)), wb.sheetnames

    stock_rows = _rows(wb["Остатки комплектующих"])
    assert any(row[1] == "Часть А" and row[2] == 20000 for row in stock_rows), stock_rows
    assert any(row[1] == "Часть Б" and row[2] == 11000 for row in stock_rows), stock_rows

    daily_rows = _rows(wb["Комплектующие по датам"])
    assert any(row[1] == "Часть А" and row[-1] == 30000 for row in daily_rows), daily_rows
    assert any(row[1] == "Часть Б" and row[-1] == 16000 for row in daily_rows), daily_rows

    plan_rows = _rows(wb["План сборки"])
    assert any(row[0] == "Изделие 1" and row[1] == 50000 and row[7] == 80000 for row in plan_rows), plan_rows
    assert any(row[0] == "Изделие 1" and row[1] == 250000 and row[7] == 480000 for row in plan_rows), plan_rows

    move_rows = _rows(wb["Собрано и отправлено"])
    assert any(row[0] == "Изделие 1" and row[1] == 5000 and row[2] == 2000 for row in move_rows), move_rows

    pdf = reporting.create_pdf_report(chat_id, "pdf отчёт за сегодня для сборки 250000", user_id=user_id)
    assert pdf.exists() and pdf.stat().st_size > 1500, pdf
    print("OK")


if __name__ == "__main__":
    main()
