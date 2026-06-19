from __future__ import annotations

import csv
import os
import sys
import tempfile
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_filetype_test_")

from app.db import init_db
from app.services import accounting, reporting, repository as repo
from app.services.parser import parse_message


def main() -> None:
    init_db()
    chat_id = -3101
    user_id = 555
    repo.set_chat_connected(chat_id, "Группа", "supergroup", True)
    ok, _ = repo.create_entity(chat_id, "component", "Деталь 1", "шт")
    assert ok
    ops, errors = parse_message(chat_id, chat_id, "Сделали Деталь 1 12")
    assert not errors, errors
    accounting.apply_operations(chat_id, chat_id, user_id, [op.to_dict() for op in ops], "Сделали Деталь 1 12")
    prefs = {"inventory": True, "period_totals": True, "daily_matrix": True, "capacity": True, "journal": True}
    repo.set_export_preferences(chat_id, user_id, prefs)

    keyboard_source = (Path(__file__).resolve().parents[1] / "app" / "keyboards.py").read_text(encoding="utf-8")
    assert "report:file:{token}:xlsx" in keyboard_source
    assert "report:file:{token}:pdf" in keyboard_source
    assert "report:file:{token}:csv" in keyboard_source
    assert "report:file:{token}:html" in keyboard_source
    assert "report:file:{token}:txt" in keyboard_source
    download_block = keyboard_source.split("def report_download_keyboard", 1)[1]
    assert "zip" not in download_block.lower()

    sections = reporting.report_sections(chat_id, "отчёт за сегодня", user_id=user_id)
    section_titles = [title for title, _, _ in sections]
    assert section_titles == ["Склад", "Итоги за период", "По датам", "Расчёт сборки", "Журнал"], section_titles

    xlsx = reporting.create_xlsx_report(chat_id, "отчёт за сегодня", user_id=user_id)
    pdf = reporting.create_pdf_report(chat_id, "отчёт за сегодня", user_id=user_id)
    csv_path = reporting.create_csv_report(chat_id, "отчёт за сегодня", user_id=user_id)
    html_path = reporting.create_html_report(chat_id, "отчёт за сегодня", user_id=user_id)
    txt = reporting.create_txt_report(chat_id, "отчёт за сегодня", user_id=user_id)

    for path in (xlsx, pdf, csv_path, html_path, txt):
        assert path.exists() and path.stat().st_size > 100, path
    assert xlsx.suffix == ".xlsx"
    assert pdf.suffix == ".pdf"
    assert csv_path.suffix == ".csv"
    assert html_path.suffix == ".html"
    assert txt.suffix == ".txt"

    workbook = load_workbook(xlsx, read_only=True)
    assert workbook.sheetnames == section_titles, workbook.sheetnames
    assert [cell.value for cell in next(workbook["Итоги за период"].iter_rows(max_row=1))] == sections[1][1]

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        csv_text = f.read()
    assert "Склад" in csv_text and "Итоги за период" in csv_text and "Расчёт сборки" in csv_text
    html_text = html_path.read_text(encoding="utf-8")
    assert "<table" in html_text and "Склад" in html_text and "Итоги за период" in html_text
    txt_text = txt.read_text(encoding="utf-8")
    assert "Склад" in txt_text and "Итоги за период" in txt_text and "Расчёт сборки" in txt_text
    assert " | " in txt_text

    universal = reporting.create_universal_report_zip(chat_id, "отчёт за сегодня", user_id=user_id)
    with ZipFile(universal) as zf:
        names = zf.namelist()
    assert any(name.endswith(".xlsx") for name in names)
    assert any(name.endswith(".pdf") for name in names)
    print("OK")


if __name__ == "__main__":
    main()
