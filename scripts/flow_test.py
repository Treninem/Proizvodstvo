from __future__ import annotations

import os
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_flow_")
os.environ.setdefault("GLOBAL_OWNER_IDS", "2097006037")

from app.db import init_db
from app.services import accounting, backups, reporting, repository as repo
from app.services.parser import looks_like_accounting, parse_message


def _apply(scope_chat_id: int, group_chat_id: int, user_id: int, text: str) -> int:
    parsed, errors = parse_message(scope_chat_id, group_chat_id, text)
    assert not errors, errors
    assert parsed, text
    return accounting.apply_operations(scope_chat_id, group_chat_id, user_id, [item.to_dict() for item in parsed], text)


def main() -> None:
    init_db()

    owner_id = 2097006037
    user_full = 7001
    user_submit = 7002
    group_a = -101001
    group_b = -101002
    private_chat = 701000

    repo.set_chat_connected(group_a, "Группа 1", "supergroup", True)
    repo.set_chat_connected(group_b, "Группа 2", "supergroup", True)
    repo.upsert_chat(private_chat, "Личный чат", "private", connected=None)

    ok, msg, account_id = repo.create_account(user_full, group_a, "Учёт 1")
    assert ok and account_id is not None, msg
    scope = repo.resolve_scope_chat_id(group_a)
    assert scope != group_a

    ok, msg = repo.create_or_set_self_job(group_a, user_full, "Участник 1", "Должность 1")
    assert ok, msg
    assert repo.user_permissions_current_context(group_a, user_full).get("setup") is True

    ok, msg = repo.create_area(group_a, "Участок 1")
    assert ok, msg
    ok, msg = repo.create_area(group_a, "Участок 2")
    assert ok, msg
    areas = repo.list_areas(group_a)
    area1 = next(a for a in areas if a.name == "Участок 1")
    area2 = next(a for a in areas if a.name == "Участок 2")
    repo.bind_chat_to_area(group_a, area1.id)
    repo.attach_chat_to_account(account_id, group_b, can_manage=False, set_active=True)
    repo.bind_chat_to_area(group_b, area2.id)

    ok, msg = repo.create_job_title(group_a, "Должность 2", {"production": True, "material": True, "energy": True, "reports": True})
    assert ok, msg
    job = repo.find_job_title(group_a, "Должность 2")
    assert job
    repo.set_worker_job(group_a, user_submit, "Участник 2", int(job["id"]))
    assert repo.user_permissions_current_context(group_a, user_submit).get("production") is True
    assert repo.user_permissions_current_context(group_a, user_submit).get("setup") is not True

    ok, msg = repo.set_active_account(private_chat, account_id, user_id=user_submit)
    assert ok, msg
    assert repo.user_permissions_current_context(private_chat, user_submit).get("material") is True

    for entity_type, name, unit in (
        ("component", "Комплектующая 1", "шт"),
        ("component", "Комплектующая 2", "шт"),
        ("product", "Изделие 1", "шт"),
        ("product", "Изделие 2", "шт"),
        ("material", "Сырьё 1", "кг"),
        ("material", "Сырьё 2", "кг"),
        ("meter", "Счётчик 1", "кВт⋅ч"),
        ("stock_item", "Позиция склада 1", "шт"),
    ):
        ok, msg = repo.create_entity(group_a, entity_type, name, unit)
        assert ok, msg

    comp1, comp2 = repo.list_entities(group_a, {"component"})
    prod1, prod2 = repo.list_entities(group_a, {"product"})
    repo.set_product_components(group_a, prod1.id, [(comp1.id, 2), (comp2.id, 1)])
    repo.set_product_components(group_a, prod2.id, [(comp1.id, 1), (comp2.id, 3)])
    assert len(repo.list_product_components(prod1.id)) == 2
    assert len(repo.list_product_components(prod2.id)) == 2

    meter = repo.list_entities(group_a, {"meter"})[0]
    repo.bind_meter_to_areas(group_a, meter.id, [area1.id, area2.id])
    assert set(repo.list_meter_area_names(meter.id)) == {"Участок 1", "Участок 2"}

    repo.add_aliases(group_a, "component", comp1.id, "к1, комп1, первая")
    repo.add_aliases(group_a, "material", repo.list_entities(group_a, {"material"})[0].id, "с1, материал 1")

    assert not looks_like_accounting("Нужно завтра сделать Комплектующая 1")
    assert looks_like_accounting("Сделали Комплектующая 1 20")
    assert looks_like_accounting("Привезли\nУчасток 1 Сырьё 1 500 кг\nСырьё 2 100 кг")

    _apply(group_a, group_a, user_submit, "Сделали Комплектующая 1 40\nКомплектующая 2 30")
    _apply(group_a, group_a, user_submit, "Привезли\nУчасток 1 Сырьё 1 500 кг\nСырьё 2 100 кг")
    _apply(group_a, group_b, user_submit, "Участок 2 Счётчик 1 1550.5")
    _apply(group_a, group_a, user_submit, "Собрано Изделие 1 5")
    _apply(group_a, group_a, user_submit, "Отправлено Изделие 1 2")

    capacity_one = reporting.build_assembly_capacity_report(group_a, "сколько можно собрать Изделие 1")
    assert "Изделие 1" in capacity_one and "Итого" in capacity_one, capacity_one
    capacity_all = reporting.build_assembly_capacity_report(group_a, "сколько можно собрать все")
    assert "Изделие 1" in capacity_all and "Изделие 2" in capacity_all, capacity_all

    report_today = reporting.build_text_report(group_a, "отчёт за сегодня")
    assert "Отчёт" in report_today and "Остатки" in report_today, report_today
    report_bad = reporting.build_text_report(group_a, "20.06.2022-12.05.2022")
    assert "не может быть позже" in report_bad, report_bad

    xlsx = reporting.create_xlsx_report(group_a, "excel отчёт за сегодня", user_id=user_submit)
    pdf = reporting.create_pdf_report(group_a, "pdf отчёт за сегодня", user_id=user_submit)
    for path in (xlsx, pdf):
        assert path.exists(), path
    workbook = load_workbook(xlsx, read_only=True)
    assert "Склад" in workbook.sheetnames and "Расчёт сборки" in workbook.sheetnames, workbook.sheetnames

    recent = accounting.list_recent_operations(group_a, group_a, user_submit, 10)
    assert recent, recent
    edit_id = int(recent[0]["id"])
    ok, msg = accounting.cancel_operation(group_a, group_a, user_submit, edit_id)
    assert ok, msg

    backup = backups.create_account_backup(group_a, user_id=user_submit)
    assert backup.exists() and backup.suffix == ".zip", backup
    assert repo.user_permissions_current_context(group_a, owner_id).get("setup") is True
    assert repo.user_permissions_current_context(group_a, owner_id).get("export") is True

    shutil.rmtree(os.environ["BOT_DATA_DIR"], ignore_errors=True)
    print("OK")


if __name__ == "__main__":
    main()
