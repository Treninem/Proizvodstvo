from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_report_quality_")

from app.db import init_db
from app.services import accounting, reporting, repository as repo
from app.services.parser import parse_message


def main() -> None:
    init_db()
    chat_id = -3501
    user_id = 777
    repo.set_chat_connected(chat_id, "Группа", "supergroup", True)
    for name in ("Комплектующая 1", "Комплектующая 2"):
        ok, _ = repo.create_entity(chat_id, "component", name, "шт")
        assert ok
    ok, _ = repo.create_entity(chat_id, "product", "Изделие 1", "шт")
    assert ok
    product = repo.list_entities(chat_id, {"product"})[0]
    components = repo.list_entities(chat_id, {"component"})
    repo.set_product_components(chat_id, product.id, [(components[0].id, 2), (components[1].id, 1)])

    messages = [
        "Сделали Комплектующая 1 20",
        "Сделали Комплектующая 2 8",
        "Собрали Изделие 1 3",
    ]
    for message in messages:
        ops, errors = parse_message(chat_id, chat_id, message)
        assert not errors, errors
        accounting.apply_operations(chat_id, chat_id, user_id, [op.to_dict() for op in ops], message)

    repo.set_export_preferences(chat_id, user_id, {
        "inventory": True,
        "period_totals": False,
        "daily_matrix": False,
        "capacity": False,
        "journal": False,
    })

    xlsx = reporting.create_xlsx_report(chat_id, "отчёт за сегодня", user_id=user_id)
    workbook = load_workbook(xlsx, read_only=False)
    assert workbook.sheetnames[0] == "Отчёт", workbook.sheetnames
    for title in ("Склад", "Итоги за период", "По датам", "Расчёт сборки", "Журнал"):
        assert title in workbook.sheetnames, workbook.sheetnames
        assert title in [cell.value for row in workbook["Отчёт"].iter_rows() for cell in row if cell.value], title

    pdf = reporting.create_pdf_report(chat_id, "отчёт за сегодня", user_id=user_id)
    data = pdf.read_bytes()
    assert pdf.exists() and pdf.stat().st_size > 1500, pdf
    assert b"/ToUnicode" in data, "PDF должен содержать карту символов для русского текста"
    print("OK")


if __name__ == "__main__":
    main()
