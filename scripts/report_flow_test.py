from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_report_test_")

from app.db import init_db
from app.services import accounting, reporting, repository as repo
from app.services.parser import parse_message


def main() -> None:
    init_db()
    keyboard_source = (Path(__file__).resolve().parents[1] / "app" / "keyboards.py").read_text(encoding="utf-8")
    banned = "Файл " + "для " + "печати"
    assert banned not in keyboard_source

    chat_id = -2001
    repo.set_chat_connected(chat_id, "Группа 1", "supergroup", True)
    ok, _ = repo.create_entity(chat_id, "component", "Деталь 1", "шт")
    assert ok
    ops, errors = parse_message(chat_id, chat_id, "Сделали Деталь 1 15")
    assert not errors, errors
    accounting.apply_operations(chat_id, chat_id, 111, [op.to_dict() for op in ops], "Сделали Деталь 1 15")

    repo.set_export_preferences(chat_id, 111, {
        "inventory": False,
        "period_totals": True,
        "daily_matrix": False,
        "capacity": False,
        "journal": False,
    })
    text = reporting.build_text_report(chat_id, "отчёт за сегодня", user_id=111)
    assert "Производство" in text, text
    assert "Остатки склада" not in text, text

    path = reporting.create_xlsx_report(chat_id, "отчёт за сегодня", user_id=111)
    assert path.exists() and path.suffix == ".xlsx", path
    workbook = load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["Итоги за период"], workbook.sheetnames
    print("OK")


if __name__ == "__main__":
    main()
