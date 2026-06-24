from __future__ import annotations

import os
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

os.environ["BOT_DATA_DIR"] = tempfile.mkdtemp(prefix="prod_bot_test_")

from app import db
from app.db import init_db
from app.services import repository as repo
from app.services.parser import is_yes, looks_like_accounting, parse_message
from app.services import accounting
from app.services import reporting
from app.services import backups
from app.services.inventory_adjustment import parse_inventory_lines
from app.services.vocabulary import synonym_counts, MIN_SYNONYMS_PER_OPERATION


def main() -> None:
    init_db()
    chat_id = -1001
    repo.set_chat_connected(chat_id, "Группа 1", "supergroup", True)
    ok, _ = repo.create_area(chat_id, "Участок 1")
    assert ok
    area = repo.list_areas(chat_id)[0]
    repo.bind_chat_to_area(chat_id, area.id)
    ok, _ = repo.create_entity(chat_id, "component", "Деталь 1", "шт")
    assert ok
    ok, _ = repo.create_entity(chat_id, "component", "Деталь 2", "шт")
    assert ok
    ok, _ = repo.create_entity(chat_id, "product", "Изделие 1", "шт")
    assert ok
    product = repo.list_entities(chat_id, {"product"})[0]
    components_for_test = repo.list_entities(chat_id, {"component"})
    repo.set_product_components(chat_id, product.id, [(components_for_test[0].id, 1), (components_for_test[1].id, 2)])
    assert len(repo.list_product_components(product.id)) == 2
    ok, _ = repo.create_job_title(chat_id, "Должность 1", {"production": True, "material": True})
    assert ok
    job = repo.find_job_title(chat_id, "Должность 1")
    repo.set_worker_job(chat_id, 888, "Работник 1", int(job["id"]))
    assert repo.worker_permissions(chat_id, 888).get("production") is True
    ok, _ = repo.create_entity(chat_id, "material", "Сырьё 1", "кг")
    assert ok
    ok, _ = repo.create_entity(chat_id, "material", "Сырьё 2", "кг")
    assert ok
    ok, _ = repo.create_entity(chat_id, "stock_item", "Позиция склада 1", "шт")
    assert ok
    ok, _ = repo.create_entity(chat_id, "meter", "Счётчик 1", "кВт⋅ч")
    assert ok
    meter = repo.list_entities(chat_id, {"meter"})[0]
    repo.bind_meter_to_areas(chat_id, meter.id, [area.id])
    assert repo.list_meter_area_names(meter.id) == ["Участок 1"]
    repo.add_aliases(chat_id, "component", repo.list_entities(chat_id, {"component"})[0].id, "д1, первая")
    assert looks_like_accounting("Производство Участок 1\nДеталь 1 300\nРасход Сырьё 1 25 кг")
    assert looks_like_accounting("Изготовили Деталь 1 10")
    assert not looks_like_accounting("Нужно завтра сделать Комплектующая 1")
    assert looks_like_accounting("Потрачено Сырьё 1 2 кг")
    assert looks_like_accounting("Свет 1555")
    assert looks_like_accounting("Оприходовали Позиция склада 1 4")
    ops, errors = parse_message(chat_id, chat_id, "Производство Участок 1\nДеталь 1 300\nДеталь 2 600\nРасход Сырьё 1 25 кг")
    assert len(ops) == 3, ops
    assert any(op.operation_type == "material_out" for op in ops)
    ops2, _ = parse_message(chat_id, chat_id, "Привезли\nУчасток 1 Сырьё 1 500 кг\nСырьё 2 300 кг")
    assert len(ops2) == 2, ops2
    assert all(op.area_id == area.id for op in ops2)
    ops3, errors3 = parse_message(chat_id, chat_id, "Приход Позиция склада 1 12")
    assert len(ops3) == 1 and ops3[0].operation_type == "stock_in", ops3
    assert not errors3
    ops4, errors4 = parse_message(chat_id, chat_id, "Уход Позиция склада 1 3")
    assert len(ops4) == 1 and ops4[0].operation_type == "stock_out", ops4
    ops5, errors5 = parse_message(chat_id, chat_id, "Счётчик 1 1356.7")
    assert len(ops5) == 1 and ops5[0].area_id == area.id, ops5
    assert not errors5, errors5

    # Складская позиция может быть общей: даже в группе участка она остаётся без участка.
    stock_item = repo.list_entities(chat_id, {"stock_item"})[0]
    ops6, errors6 = parse_message(chat_id, chat_id, "Приход Позиция склада 1 12")
    assert len(ops6) == 1 and ops6[0].operation_type == "stock_in", ops6
    assert ops6[0].area_id is None, ops6
    assert not errors6, errors6

    # Складскую позицию можно привязать к одному участку.
    repo.bind_stock_item_to_areas(chat_id, stock_item.id, [area.id])
    ops7, errors7 = parse_message(chat_id, chat_id, "Уход Позиция склада 1 3")
    assert len(ops7) == 1 and ops7[0].area_id == area.id, ops7
    assert not errors7, errors7

    # Если счётчик привязан к нескольким участкам, но участок понятен из группы, распределение не спрашивается.
    ok, _ = repo.create_area(chat_id, "Участок 2")
    assert ok
    area2 = [a for a in repo.list_areas(chat_id) if a.name == "Участок 2"][0]
    repo.bind_meter_to_areas(chat_id, meter.id, [area.id, area2.id])
    ops8, errors8 = parse_message(chat_id, chat_id, "Счётчик 1 1400")
    assert len(ops8) == 1 and ops8[0].area_id == area.id, ops8
    assert not errors8, errors8

    # Если указан только участок и показание э/э, бот подставляет закреплённый прибор учёта.
    ops9, errors9 = parse_message(chat_id, chat_id, "Участок 2 ээ 1500")
    assert len(ops9) == 1 and ops9[0].entity_id == meter.id and ops9[0].area_id == area2.id, ops9
    assert not errors9, errors9


    ops10, errors10 = parse_message(chat_id, chat_id, "Изготовили первая 10")
    assert len(ops10) == 1 and ops10[0].operation_type == "production", ops10
    assert ops10[0].entity_name == "Деталь 1", ops10
    ops11, errors11 = parse_message(chat_id, chat_id, "Потрачено Сырьё 1 2 кг")
    assert len(ops11) == 1 and ops11[0].operation_type == "material_out", ops11
    ops12, errors12 = parse_message(chat_id, chat_id, "Оприходовали Позиция склада 1 4")
    assert len(ops12) == 1 and ops12[0].operation_type == "stock_in", ops12
    assert not errors12, errors12

    counts = synonym_counts()
    assert counts, counts
    assert all(n >= MIN_SYNONYMS_PER_OPERATION for n in counts.values()), counts

    # После подтверждения бот запоминает локальный жаргон именно в словарь этого чата.
    ops10_payload = [op.to_dict() for op in ops10]
    saved = accounting.apply_operations(chat_id, chat_id, 777, ops10_payload, "Изготовили первая 10")
    assert saved == 1
    candidates = repo.list_alias_candidates(chat_id)
    assert any(c["source"] == "lexicon" and c["key"] == "первая" and c["name"] == "Деталь 1" for c in candidates), candidates

    # Сборка списывает комплектующие и добавляет готовое изделие по составу.
    accounting.apply_operations(chat_id, chat_id, 777, [op.to_dict() for op in parse_message(chat_id, chat_id, "Деталь 2 20")[0]], "Деталь 2 20")
    assembly_ops, assembly_errors = parse_message(chat_id, chat_id, "Собрано Изделие 1 5")
    assert len(assembly_ops) == 1 and assembly_ops[0].operation_type == "assembly", assembly_ops
    accounting.apply_operations(chat_id, chat_id, 777, [op.to_dict() for op in assembly_ops], "Собрано Изделие 1 5")
    capacity_report = reporting.build_assembly_capacity_report(chat_id, "сколько можно собрать Изделие 1")
    assert "Изделие 1" in capacity_report and "Итого" in capacity_report, capacity_report
    all_capacity_report = reporting.build_assembly_capacity_report(chat_id, "сколько можно собрать все")
    assert "Расчёт сборки" in all_capacity_report and "Изделие 1" in all_capacity_report, all_capacity_report

    # Отчёты и файлы.
    accounting.apply_operations(chat_id, chat_id, 777, [op.to_dict() for op in ops2], "Привезли")
    accounting.apply_operations(chat_id, chat_id, 777, [op.to_dict() for op in ops3], "Приход Позиция склада 1 12")
    text_report = reporting.build_text_report(chat_id, "отчёт за сегодня")
    assert "Отчёт" in text_report and "Склад" in text_report, text_report
    xlsx_path = reporting.create_xlsx_report(chat_id, "excel отчёт за сегодня")
    pdf_path = reporting.create_pdf_report(chat_id, "pdf отчёт за сегодня")
    csv_path = reporting.create_csv_report(chat_id, "csv отчёт за сегодня")
    html_path = reporting.create_html_report(chat_id, "html отчёт за сегодня")
    txt_path = reporting.create_txt_report(chat_id, "txt отчёт за сегодня")
    zip_path = reporting.create_universal_report_zip(chat_id, "универсальный файл за сегодня")
    assert xlsx_path.exists() and xlsx_path.suffix == ".xlsx", xlsx_path
    assert pdf_path.exists() and pdf_path.suffix == ".pdf", pdf_path
    assert csv_path.exists() and csv_path.suffix == ".csv", csv_path
    assert html_path.exists() and html_path.suffix == ".html", html_path
    assert txt_path.exists() and txt_path.suffix == ".txt", txt_path
    assert zip_path.exists() and zip_path.suffix == ".zip", zip_path
    with zipfile.ZipFile(zip_path) as zf:
        names = set(zf.namelist())
    assert any(name.endswith(".xlsx") for name in names), names
    assert any(name.endswith(".pdf") for name in names), names
    assert any(name.endswith(".csv") for name in names), names
    assert any(name.endswith(".html") for name in names), names
    assert any(name.endswith(".txt") for name in names), names

    first_component = repo.list_entities(chat_id, {"component"})[0]


    # Инвентаризация выставляет фактический остаток через подтверждаемую правку.
    inv_ops, inv_errors = parse_inventory_lines(chat_id, chat_id, "Инвентаризация\nДеталь 1 100")
    assert len(inv_ops) == 1 and inv_ops[0]["operation_type"] == "inventory_adjust", inv_ops
    assert not inv_errors, inv_errors
    accounting.apply_operations(chat_id, chat_id, 777, inv_ops, "Инвентаризация")
    fact_qty = repo.inventory_quantity(chat_id, "component", first_component.id, "шт", None)
    assert abs(fact_qty - 100) < 0.0001, fact_qty
    db.execute(
        """
        INSERT INTO operations(chat_id,group_chat_id,area_id,user_id,operation_type,entity_type,entity_id,quantity,unit,raw_text,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?)
        """,
        (chat_id, chat_id, None, 777, "production", "component", first_component.id, 42, "шт", "Проверка периода", "2022-05-12 10:00:00"),
    )
    db.execute(
        """
        INSERT INTO operations(chat_id,group_chat_id,area_id,user_id,operation_type,entity_type,entity_id,quantity,unit,raw_text,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?)
        """,
        (chat_id, chat_id, None, 777, "shipment", "product", product.id, 9, "шт", "Вне периода", "2023-05-12 10:00:00"),
    )
    range_report = reporting.build_text_report(chat_id, "12.05.2022-20.06.2022")
    assert "12.05.2022" in range_report and "42" in range_report, range_report
    one_day_report = reporting.build_text_report(chat_id, "отчёт 12.05.2022")
    assert "за 12.05.2022" in one_day_report and "42" in one_day_report, one_day_report
    invalid_report = reporting.build_text_report(chat_id, "11.05.2023-20.06.2022")
    assert "не может быть позже" in invalid_report, invalid_report
    too_long_report = reporting.build_text_report(chat_id, "01.01.1900-02.01.2001")
    assert "100 лет" in too_long_report, too_long_report
    range_xlsx = reporting.create_xlsx_report(chat_id, "excel 12.05.2022-20.06.2022")
    assert range_xlsx.exists() and range_xlsx.suffix == ".xlsx", range_xlsx
    workbook = load_workbook(range_xlsx, read_only=True)
    assert "По датам" in workbook.sheetnames and "Расчёт сборки" in workbook.sheetnames, workbook.sheetnames
    repo.set_export_preference(chat_id, 777, "journal", False)
    custom_xlsx = reporting.create_xlsx_report(chat_id, "excel 12.05.2022-20.06.2022", user_id=777)
    custom_workbook = load_workbook(custom_xlsx, read_only=True)
    assert "Журнал" in custom_workbook.sheetnames and "Склад" in custom_workbook.sheetnames, custom_workbook.sheetnames
    repo.set_export_preference(chat_id, 777, "journal", True)


    # Исправления и отмена записей не ломают склад: создаётся обратная запись и журнал правки.
    before_rows = accounting.list_recent_operations(chat_id, group_chat_id=chat_id, user_id=777, limit=5)
    assert before_rows, before_rows
    last_id = int(before_rows[0]["id"])
    ok_cancel, cancel_msg = accounting.cancel_operation(chat_id, chat_id, 777, last_id)
    assert ok_cancel, cancel_msg
    ok_cancel_again, _ = accounting.cancel_operation(chat_id, chat_id, 777, last_id)
    assert not ok_cancel_again

    prod_ops, _ = parse_message(chat_id, chat_id, "Изготовили Деталь 1 30")
    accounting.apply_operations(chat_id, chat_id, 777, [op.to_dict() for op in prod_ops], "Изготовили Деталь 1 30")
    last_prod_id = accounting.last_editable_operation_id(chat_id, chat_id, 777)
    assert last_prod_id is not None
    ok_change, change_msg = accounting.change_operation_quantity(chat_id, chat_id, 777, int(last_prod_id), 12)
    assert ok_change, change_msg
    recent_text = accounting.format_recent_operations(accounting.list_recent_operations(chat_id, chat_id, 777, 3))
    assert "Последние записи" in recent_text and "№" in recent_text, recent_text


    account_backup = backups.create_account_backup(chat_id, user_id=777)
    assert account_backup.exists() and account_backup.suffix == ".zip", account_backup
    with zipfile.ZipFile(account_backup) as zf:
        backup_names = set(zf.namelist())
    assert any(name.endswith(".json") for name in backup_names), backup_names
    backup_list_text = backups.format_backup_list()
    assert "коп" in backup_list_text.lower() or "kopiya" in backup_list_text.lower(), backup_list_text
    full_backup = backups.create_full_backup()
    assert full_backup.exists() and full_backup.suffix == ".zip", full_backup

    stats = repo.owner_global_stats()
    assert stats["total_chats"] >= 1, stats
    assert stats["connected_chats"] >= 1, stats
    chats = repo.owner_list_chats()
    assert any(c["chat_id"] == chat_id for c in chats), chats
    chat_report = repo.owner_chat_report(chat_id)
    assert "Группа 1" in chat_report and "Операций" in chat_report, chat_report

    assert is_yes("да") and is_yes("ок")

    account_chat = -2002
    repo.set_chat_connected(account_chat, "Группа 2", "supergroup", True)
    ok, msg, account_id = repo.create_account(777, account_chat, "Учёт 1")
    assert ok, msg
    assert account_id is not None
    assert repo.resolve_scope_chat_id(account_chat) != account_chat
    ok, _ = repo.create_area(account_chat, "Участок учёта")
    assert ok
    assert repo.list_areas(account_chat)[0].name == "Участок учёта"
    repo.set_active_account(account_chat, account_id, user_id=777)
    assert repo.get_active_account(account_chat).name == "Учёт 1"

    ok, msg = repo.create_or_set_self_job(account_chat, 777, "Владелец", "Должность полного доступа")
    assert ok, msg
    assert repo.visible_job_name(account_chat, 777) == "Должность полного доступа"
    assert repo.user_permissions_current_context(account_chat, 777).get("setup") is True

    ok, _ = repo.create_job_title(account_chat, "Должность сдачи", {"production": True})
    assert ok
    small_job = repo.find_job_title(account_chat, "Должность сдачи")
    repo.set_worker_job(account_chat, 999, "Участник 2", int(small_job["id"]))
    assert repo.visible_job_name(account_chat, 999) == "Должность сдачи"
    assert repo.user_permissions_current_context(account_chat, 999).get("production") is True
    assert not repo.user_permissions_current_context(account_chat, 999).get("setup")
    accessible = repo.list_accounts_for_user(999)
    assert any(a.id == account_id for a in accessible), accessible
    private_chat = 99999
    repo.upsert_chat(private_chat, "Личный чат", "private", connected=None)
    ok, msg = repo.set_active_account(private_chat, account_id, user_id=999)
    assert ok, msg
    assert repo.user_permissions_current_context(private_chat, 999).get("production") is True
    ok, msg, common_id = repo.create_account(777, account_chat, "Общий учёт", is_general=True)
    assert ok and common_id is not None, msg
    repo.attach_chat_to_account(common_id, chat_id, can_manage=True, set_active=True)
    assert repo.get_active_account(chat_id).name == "Общий учёт"
    stats2 = repo.owner_global_stats()
    assert stats2.get("account_users", 0) >= 2, stats2

    shutil.rmtree(os.environ["BOT_DATA_DIR"], ignore_errors=True)

    print("OK")


if __name__ == "__main__":
    main()
