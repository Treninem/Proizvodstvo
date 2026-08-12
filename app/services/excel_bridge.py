from __future__ import annotations
import io, json, uuid, re
from datetime import date, datetime
from typing import Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .. import db
from . import repository as repo, accounting

MAX_BYTES=15*1024*1024
METRICS={
    'приход':'stock_in','поступление':'stock_in','пр-во':'production','производство':'production','выпуск':'production',
    'расход':'stock_out','уход':'stock_out','остаток':'balance'
}

def _norm(v:Any)->str:
    return re.sub(r'\s+',' ',str(v or '').strip().lower()).replace('ё','е')

def _metric(text:str)->str|None:
    n=_norm(text)
    for k,v in METRICS.items():
        if k in n:return v
    return None

def _entity_type(title:str)->str:
    n=_norm(title)
    if 'сыр' in n or 'материал' in n:return 'material'
    if 'детал' in n or 'комплект' in n:return 'component'
    if 'готов' in n or 'издел' in n or 'продук' in n:return 'product'
    return 'stock_item'

def _num(v:Any)->float|None:
    if isinstance(v,(int,float)) and not isinstance(v,bool):return float(v)
    t=str(v or '').strip().replace(' ','').replace(',','.')
    if not t:return None
    try:return float(t)
    except Exception:return None

def _date_text(v:Any)->str:
    if isinstance(v,datetime):return v.date().isoformat()
    if isinstance(v,date):return v.isoformat()
    t=str(v or '').strip()
    for fmt in ('%d.%m.%Y','%d.%m.%y','%Y-%m-%d'):
        try:return datetime.strptime(t,fmt).date().isoformat()
        except Exception:pass
    return ''

def _merged_value(ws,row:int,col:int):
    val=ws.cell(row,col).value
    if val not in (None,''):return val
    for rng in ws.merged_cells.ranges:
        if rng.min_row<=row<=rng.max_row and rng.min_col<=col<=rng.max_col:
            return ws.cell(rng.min_row,rng.min_col).value
    return val

def _find_header(ws)->int:
    best=(0,1)
    for r in range(1,min(ws.max_row,25)+1):
        vals=[_norm(_merged_value(ws,r,c)) for c in range(1,min(ws.max_column,40)+1)]
        score=sum(2 for v in vals if _metric(v))+sum(1 for v in vals if any(x in v for x in ('деталь','сырье','сырьё','позиция','наименование','дата','недел')))
        if score>best[0]:best=(score,r)
    return best[1]

def _find_identity_columns(ws,header:int)->tuple[int|None,int|None,int|None]:
    item=date_col=week_col=None
    for c in range(1,ws.max_column+1):
        # Use only values physically stored in the cell here. A large merged sheet title
        # (for example «СЫРЬЁ» across the whole table) must not make column A look
        # like the item-name column.
        texts=' '.join(_norm(ws.cell(r,c).value) for r in range(max(1,header-3),header+1) if ws.cell(r,c).value not in (None,''))
        is_date='дата' in texts
        is_week=('недел' in texts or '№ недели' in texts)
        if item is None and not is_date and not is_week and any(k in texts for k in ('деталь','позиция','наименование','сырье','сырьё','изделие')):item=c
        if date_col is None and is_date:date_col=c
        if week_col is None and is_week:week_col=c
    if item is None:
        # fallback: first text-heavy column after date/week
        for c in range(1,ws.max_column+1):
            sample=[str(ws.cell(r,c).value or '').strip() for r in range(header+1,min(ws.max_row,header+15)+1)]
            if sum(1 for x in sample if x and _num(x) is None)>=3:item=c;break
    return item,date_col,week_col

def _location_for_col(ws,header:int,col:int)->str:
    parts=[]
    for r in range(max(1,header-3),header+1):
        text=str(_merged_value(ws,r,col) or '').strip()
        if text and not _metric(text) and _norm(text) not in {'локация','остаток итого','итого','сырье','сырьё','детали','деталь','изделия','изделие','продукция','готовая продукция'}:
            parts.append(text)
    # nearest higher merged group usually has location name
    uniq=[]
    for p in parts:
        if _norm(p) not in {_norm(x) for x in uniq}:uniq.append(p)
    return ' / '.join(uniq[-2:])[:180] or 'Без площадки'

def analyze_bytes(chat_id:int,user_id:int,data:bytes,file_name:str)->dict:
    scope=repo.resolve_scope_chat_id(chat_id)
    if not repo.is_tenant_admin(scope,user_id):raise PermissionError('Импорт Excel доступен владельцу/администратору организации.')
    if not data or len(data)>MAX_BYTES:raise ValueError('Файл пустой или больше 15 МБ.')
    try:wb=load_workbook(io.BytesIO(data),data_only=True,read_only=False)
    except Exception as exc:raise ValueError(f'Не удалось открыть XLSX: {exc}')
    candidates=[]
    for ws in wb.worksheets:
        if ws.max_row<2 or ws.max_column<2:continue
        header=_find_header(ws);item_col,date_col,week_col=_find_identity_columns(ws,header)
        metric_cols=[]
        for c in range(1,ws.max_column+1):
            m=None;source=''
            for r in range(max(1,header-1),header+2):
                text=str(_merged_value(ws,r,c) or '')
                if _metric(text):m=_metric(text);source=text;break
            if m:metric_cols.append({'col':c,'metric':m,'source_header':source,'location':_location_for_col(ws,header,c)})
        if not item_col or not metric_cols:continue
        et=_entity_type(f'{ws.title} {file_name}')
        rows=[];warnings=[]
        default_unit='кг' if et=='material' else 'шт'
        if et=='material':warnings.append('Единица измерения в таблице не найдена: для новых позиций сырья предполагается «кг». Для уже существующих позиций используется их единица из справочника.')
        last_name=''
        for r in range(header+1,min(ws.max_row,5000)+1):
            raw_name=ws.cell(r,item_col).value
            name=str(raw_name or '').strip()
            if not name:continue
            if _num(name) is not None:continue
            last_name=name
            d=_date_text(ws.cell(r,date_col).value) if date_col else ''
            week=str(ws.cell(r,week_col).value or '').strip() if week_col else ''
            for mc in metric_cols:
                qty=_num(ws.cell(r,mc['col']).value)
                if qty is None:continue
                # zeros are retained only for explicit balances; zero movement adds no value.
                if abs(qty)<1e-12 and mc['metric']!='balance':continue
                rows.append({'sheet':ws.title,'row':r,'source_date':d,'week':week,'entity_type':et,'entity_name':name,'metric':mc['metric'],'location_name':mc['location'],'quantity':qty,'unit':default_unit,'source_column':mc['col']})
        if rows:candidates.append((len(rows),ws,header,item_col,date_col,week_col,metric_cols,et,rows,warnings))
    if not candidates:raise ValueError('Не удалось уверенно распознать колонки приход/производство/расход/остаток. Файл не изменён.')
    _,ws,header,item_col,date_col,week_col,metric_cols,et,rows,warnings=max(candidates,key=lambda x:x[0])
    batch=uuid.uuid4().hex
    mapping={'sheet':ws.title,'header_row':header,'item_column':item_col,'date_column':date_col,'week_column':week_col,'metric_columns':metric_cols,'entity_type':et}
    preview={'batch_id':batch,'file_name':str(file_name or '')[:255],'sheet':ws.title,'header_row':header,'entity_type':et,'total_rows':len(rows),'rows':rows,'mapping':mapping,'warnings':warnings}
    db.execute('''INSERT INTO excel_import_batches(batch_id,chat_id,file_name,status,created_by,preview_json) VALUES(?,?,?,'preview',?,?)''',(batch,scope,str(file_name or '')[:255],int(user_id),json.dumps(preview,ensure_ascii=False)))
    repo.tenant_audit(scope,user_id,'excel_preview','excel_import',batch,f'{file_name}; rows={len(rows)}')
    return preview

def get_preview(chat_id:int,user_id:int,batch_id:str)->dict:
    scope=repo.resolve_scope_chat_id(chat_id)
    if not repo.is_tenant_admin(scope,user_id):raise PermissionError('Нет доступа.')
    row=db.fetchone('SELECT * FROM excel_import_batches WHERE batch_id=? AND chat_id=?',(str(batch_id),scope))
    if not row:raise ValueError('Предпросмотр не найден.')
    result=json.loads(row['preview_json'] or '{}');result['status']=row['status'];return result

def _entity(scope:int,etype:str,name:str,create_missing:bool,default_unit:str='шт'):
    key=repo.normalize_key(name)
    row=db.fetchone('SELECT id,name,default_unit,entity_type FROM entities WHERE chat_id=? AND normalized=? AND is_archived=0',(scope,key))
    if row:return dict(row)
    if not create_missing:return None
    ok,_=repo.create_entity(scope,etype,name,str(default_unit or ('кг' if etype=='material' else 'шт'))[:30])
    if not ok:return None
    row=db.fetchone('SELECT id,name,default_unit,entity_type FROM entities WHERE chat_id=? AND normalized=? AND is_archived=0',(scope,key));return dict(row) if row else None

def _area(scope:int,name:str,create_missing:bool):
    clean=str(name or '').strip()
    if not clean or _norm(clean) in {'без площадки','локация'}:return None
    # If a composite header uses "A / приход", keep the leading physical label.
    clean=re.sub(r'\s*/\s*(приход|пр-во|производство|расход|остаток)\s*$','',clean,flags=re.I).strip()
    key=repo.normalize_key(clean)
    row=db.fetchone('SELECT id,name FROM areas WHERE chat_id=? AND normalized=? AND is_archived=0',(scope,key))
    if row:return dict(row)
    loc=db.fetchone('SELECT area_id,name FROM storage_locations WHERE chat_id=? AND normalized=? AND is_archived=0',(scope,key))
    if loc and loc['area_id'] is not None:
        area=db.fetchone('SELECT id,name FROM areas WHERE id=? AND chat_id=? AND is_archived=0',(int(loc['area_id']),scope))
        if area:return dict(area)
    site=db.fetchone('SELECT id FROM company_sites WHERE chat_id=? AND normalized=? AND is_archived=0',(scope,key))
    if site:
        site_areas=db.fetchall('SELECT id,name FROM areas WHERE chat_id=? AND site_id=? AND is_archived=0 ORDER BY id',(scope,int(site['id'])))
        if len(site_areas)==1:return dict(site_areas[0])
    if not create_missing:return None
    ok,_=repo.create_area(scope,clean)
    if not ok:return None
    row=db.fetchone('SELECT id,name FROM areas WHERE chat_id=? AND normalized=? AND is_archived=0',(scope,key));return dict(row) if row else None

def confirm_import(chat_id:int,user_id:int,batch_id:str,create_missing:bool=True)->dict:
    scope=repo.resolve_scope_chat_id(chat_id)
    if not repo.is_tenant_admin(scope,user_id):raise PermissionError('Нет права подтверждать импорт.')
    row=db.fetchone('SELECT * FROM excel_import_batches WHERE batch_id=? AND chat_id=?',(str(batch_id),scope))
    if not row:raise ValueError('Импорт не найден.')
    current_status=str(row['status'] or '')
    if current_status not in {'preview','processing'}:raise ValueError('Этот импорт уже обработан.')
    # Первый confirm атомарно переводит batch в processing. Если процесс оборвался после
    # части строк, повторный confirm продолжит его безопасно: каждая строка имеет
    # детерминированный client_request_id и уже применённые операции не дублируются.
    if current_status=='preview':
        with db.connect() as conn:
            cur=conn.execute("UPDATE excel_import_batches SET status='processing' WHERE batch_id=? AND chat_id=? AND status='preview'",(str(batch_id),scope))
            conn.commit()
            if int(cur.rowcount or 0)!=1:
                row=db.fetchone('SELECT * FROM excel_import_batches WHERE batch_id=? AND chat_id=?',(str(batch_id),scope))
                if not row or str(row['status'] or '')!='processing':raise ValueError('Этот импорт уже обработан.')
    preview=json.loads(row['preview_json'] or '{}');applied=0;skipped=[];ops=[]
    for rec in preview.get('rows',[]):
        try:
            request_id=f"excel:{batch_id}:{int(rec.get('row') or 0)}:{int(rec.get('source_column') or 0)}:{str(rec.get('metric') or '')}"[:120]
            existing=db.fetchone("SELECT id FROM operations WHERE chat_id=? AND user_id=? AND client_request_id=?",(scope,int(user_id),request_id))
            if existing:
                ops.append(int(existing['id']));applied+=1;continue
            ent=_entity(scope,str(rec.get('entity_type') or preview.get('entity_type') or 'stock_item'),str(rec.get('entity_name') or ''),create_missing,str(rec.get('unit') or 'шт'))
            if not ent:raise ValueError('позиция не найдена')
            area=_area(scope,str(rec.get('location_name') or ''),create_missing);area_id=int(area['id']) if area else None
            qty=float(rec.get('quantity') or 0);metric=rec.get('metric');unit=str(rec.get('unit') or ent.get('default_unit') or 'шт')
            if metric=='balance':
                current=repo.inventory_quantity(scope,ent['entity_type'],int(ent['id']),unit,area_id);delta=qty-current
                if abs(delta)<1e-12:continue
                op_type='inventory_adjust';op_qty=delta
            else:
                op_type={'stock_in':'stock_in','stock_out':'stock_out','production':'production'}.get(metric)
                if not op_type:continue
                op_qty=qty
            op={'operation_type':op_type,'entity_type':ent['entity_type'],'entity_id':int(ent['id']),'quantity':op_qty,'unit':unit,'area_id':area_id,'source_channel':'excel','client_request_id':request_id}
            oid=accounting.record_internal_operation(scope,scope,user_id,op,raw_text=f"Excel импорт {preview.get('file_name','')} · строка {rec.get('row')}")
            source_date=str(rec.get('source_date') or '')
            if source_date:
                db.execute("UPDATE operations SET created_at=? WHERE id=? AND chat_id=?",(source_date+' 12:00:00',oid,scope))
            ops.append(oid);applied+=1
        except Exception as exc:
            skipped.append({'row':rec.get('row'),'entity_name':rec.get('entity_name'),'reason':str(exc)[:200]})
    status='applied' if applied else 'failed'
    result={'applied':applied,'total':len(preview.get('rows',[])),'skipped':skipped,'operation_ids':ops}
    db.execute('UPDATE excel_import_batches SET status=?,confirmed_by=?,confirmed_at=CURRENT_TIMESTAMP,result_json=? WHERE batch_id=? AND chat_id=?',(status,int(user_id),json.dumps(result,ensure_ascii=False),str(batch_id),scope))
    repo.tenant_audit(scope,user_id,'excel_apply','excel_import',batch_id,f'applied={applied}; skipped={len(skipped)}',severity='warning' if skipped else 'info')
    return result

def cancel_import(chat_id:int,user_id:int,batch_id:str)->None:
    scope=repo.resolve_scope_chat_id(chat_id)
    if not repo.is_tenant_admin(scope,user_id):raise PermissionError('Нет доступа.')
    db.execute("UPDATE excel_import_batches SET status='cancelled',confirmed_by=?,confirmed_at=CURRENT_TIMESTAMP WHERE batch_id=? AND chat_id=? AND status='preview'",(int(user_id),str(batch_id),scope))
    repo.tenant_audit(scope,user_id,'excel_cancel','excel_import',batch_id,'cancelled')

def build_location_ledger_xlsx(chat_id:int,user_id:int,entity_type:str='component',start_date:str='',end_date:str='')->bytes:
    scope=repo.resolve_scope_chat_id(chat_id)
    account=repo.get_account_by_scope(scope)
    if not account or not repo.user_has_account_access(account.id,user_id):raise PermissionError('Нет доступа.')
    allowed={'material','component','product','stock_item'}
    if entity_type not in allowed:entity_type='component'
    entities=repo.list_entities(scope,[entity_type]);areas=repo.list_areas(scope)
    area_site_meta={int(r['id']):dict(r) for r in db.fetchall('''SELECT a.id,a.name AS area_name,s.name AS site_name,s.settlement FROM areas a LEFT JOIN company_sites s ON s.id=a.site_id WHERE a.chat_id=? AND a.is_archived=0''',(scope,))}
    wb=Workbook();ws=wb.active;ws.title={'material':'Сырьё','component':'Детали','product':'Продукция','stock_item':'Склад'}[entity_type]
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(4,3+4*max(1,len(areas))+1));ws.cell(1,1).value=ws.title.upper();ws.cell(1,1).font=Font(bold=True,size=14);ws.cell(1,1).alignment=Alignment(horizontal='center')
    ws.cell(2,1,'№ недели');ws.cell(2,2,'Дата');ws.cell(2,3,{'material':'Сырьё','component':'Деталь','product':'Изделие','stock_item':'Позиция'}[entity_type])
    col=4
    use_areas=areas or [None]
    for a in use_areas:
        meta=area_site_meta.get(int(a.id),{}) if a else {};name=' → '.join(x for x in (str(meta.get('settlement') or '').strip(),str(meta.get('site_name') or '').strip(),a.name if a else 'Без площадки') if x);ws.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+3);ws.cell(2,col,name or (a.name if a else 'Без площадки'))
        for j,v in enumerate(('приход','пр-во','расход','остаток')):ws.cell(3,col+j,v)
        col+=4
    ws.cell(2,col,'ИТОГО ОСТАТОК');ws.merge_cells(start_row=2,start_column=col,end_row=3,end_column=col)
    for c in range(1,col+1):ws.cell(2,c).font=Font(bold=True);ws.cell(2,c).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True);ws.cell(3,c).font=Font(bold=True);ws.cell(3,c).alignment=Alignment(horizontal='center')
    today=date.today();rowno=4;thin=Side(style='thin',color='808080')
    for ent in entities:
        ws.cell(rowno,1,today.isocalendar().week);ws.cell(rowno,2,today);ws.cell(rowno,2).number_format='dd.mm.yyyy';ws.cell(rowno,3,ent.name);total=0.0;colx=4
        for a in use_areas:
            aid=a.id if a else None;params=[scope,ent.id,aid,aid];where="o.chat_id=? AND o.entity_id=? AND ((o.area_id IS NULL AND ? IS NULL) OR o.area_id=?)"
            if start_date:where+=' AND date(o.created_at)>=date(?)';params.append(start_date)
            if end_date:where+=' AND date(o.created_at)<=date(?)';params.append(end_date)
            ops=db.fetchall(f"SELECT operation_type,COALESCE(SUM(quantity),0) q FROM operations o WHERE {where} GROUP BY operation_type",tuple(params));sums={str(x['operation_type']):float(x['q'] or 0) for x in ops}
            incoming=sums.get('stock_in',0)+sums.get('material_in',0)+sums.get('return',0);produced=sums.get('production',0)+sums.get('assembly',0);outgoing=sums.get('stock_out',0)+sums.get('material_out',0)+sums.get('shipment',0)+sums.get('shipment_client',0)+sums.get('shipment_fulfillment',0)+sums.get('write_off',0)
            remain=repo.inventory_quantity(scope,entity_type,ent.id,ent.default_unit,aid);total+=remain
            for j,v in enumerate((incoming,produced,outgoing,remain)):ws.cell(rowno,colx+j,v if abs(v)>1e-12 else 0)
            colx+=4
        ws.cell(rowno,colx,total);rowno+=1
    for rr in ws.iter_rows(min_row=2,max_row=max(3,rowno-1),min_col=1,max_col=col):
        for cell in rr:cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.column_dimensions['A'].width=10;ws.column_dimensions['B'].width=13;ws.column_dimensions['C'].width=30
    for c in range(4,col+1):ws.column_dimensions[get_column_letter(c)].width=12
    out=io.BytesIO();wb.save(out);return out.getvalue()
