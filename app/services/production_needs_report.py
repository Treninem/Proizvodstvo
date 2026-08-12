from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .. import db
from . import repository as repo
from . import reporting


def _scope(chat_id:int)->int:return repo.resolve_scope_chat_id(int(chat_id))

def _usable_stock(scope:int,entity_id:int,area_id:int|None)->float:
    if area_id:
        row=db.fetchone("SELECT COALESCE(SUM(quantity),0) AS q FROM inventory WHERE chat_id=? AND entity_id=? AND area_id=?",(scope,entity_id,area_id))
        quarantined=db.fetchone("""SELECT COALESCE(SUM(li.quantity),0) AS q FROM lot_inventory li JOIN production_lots l ON l.id=li.lot_id WHERE l.chat_id=? AND l.entity_id=? AND li.area_id=? AND l.status IN ('quarantine','rejected')""",(scope,entity_id,area_id))
    else:
        row=db.fetchone("SELECT COALESCE(SUM(quantity),0) AS q FROM inventory WHERE chat_id=? AND entity_id=?",(scope,entity_id));quarantined=db.fetchone("""SELECT COALESCE(SUM(li.quantity),0) AS q FROM lot_inventory li JOIN production_lots l ON l.id=li.lot_id WHERE l.chat_id=? AND l.entity_id=? AND l.status IN ('quarantine','rejected')""",(scope,entity_id))
    return float(row["q"] or 0)-float(quarantined["q"] or 0)

def build_rows(chat_id:int,user_id:int,*,start_date:str|None=None,end_date:str|None=None)->list[dict[str,Any]]:
    scope=_scope(chat_id);start=start_date or datetime.now().strftime("%Y-%m-%d");end=end_date or (datetime.now()+timedelta(days=7)).strftime("%Y-%m-%d")
    tasks=db.fetchall("""SELECT t.*,e.name AS output_name,d.name AS department_name,a.name AS area_name FROM production_tasks t JOIN entities e ON e.id=t.entity_id AND e.chat_id=t.chat_id JOIN departments d ON d.id=t.department_id AND d.chat_id=t.chat_id LEFT JOIN areas a ON a.id=t.area_id AND a.chat_id=t.chat_id WHERE t.chat_id=? AND t.status<>'cancelled' AND date(COALESCE(t.due_at,t.created_at)) BETWEEN date(?) AND date(?) ORDER BY COALESCE(t.due_at,t.created_at),t.id""",(scope,start,end))
    visible=repo.visible_entity_ids_for_user(scope,user_id); visible_set=None if visible is None else {int(x) for x in visible}
    rows=[]
    for tr in tasks:
        t=dict(tr)
        if not repo.is_tenant_admin(scope, user_id) and int(repo.department_actor_level(int(t["department_id"]),user_id) or 0)<10:continue
        comps=db.fetchall("SELECT pc.component_id,pc.quantity,e.name,e.default_unit FROM product_components pc JOIN entities p ON p.id=pc.product_id JOIN entities e ON e.id=pc.component_id AND e.chat_id=p.chat_id WHERE pc.product_id=? AND p.chat_id=?",(int(t["entity_id"]),scope))
        requirements=[dict(c) for c in comps]
        # Also include direct yield rule for materials tied to this planned output.
        for rr in db.fetchall("""SELECT r.entity_id,r.yield_input_qty,r.yield_output_qty,e.name,e.default_unit FROM stock_alert_rules r JOIN entities e ON e.id=r.entity_id AND e.chat_id=r.chat_id WHERE r.chat_id=? AND r.is_enabled=1 AND r.yield_output_entity_id=? AND r.yield_input_qty>0 AND r.yield_output_qty>0""",(scope,int(t["entity_id"]))):
            r=dict(rr);requirements.append({"component_id":int(r["entity_id"]),"quantity":float(r["yield_input_qty"])/float(r["yield_output_qty"]),"name":r["name"],"default_unit":r["default_unit"]})
        if not requirements:
            rows.append({"task_id":int(t["id"]),"department":t["department_name"],"output":t["output_name"],"plan":float(t["target_quantity"] or 0),"actual":float(t["actual_quantity"] or 0),"input":"—","required":0.0,"available":0.0,"open_replenishment":0.0,"shortage":0.0,"unit":t.get("unit") or "шт","due_at":t.get("due_at") or ""})
            continue
        for c in requirements:
            eid=int(c["component_id"])
            if visible_set is not None and eid not in visible_set:continue
            required=max(0.0,float(t["target_quantity"] or 0)*float(c["quantity"] or 0));available=_usable_stock(scope,eid,int(t["area_id"]) if t.get("area_id") else None)
            req=db.fetchone("""SELECT COALESCE(SUM(requested_quantity),0) AS q FROM replenishment_requests WHERE chat_id=? AND entity_id=? AND (area_id=? OR area_id IS NULL) AND status IN ('requested','approved','ordered','partial')""",(scope,eid,t.get("area_id")));open_qty=float(req["q"] or 0) if req else 0.0
            rows.append({"task_id":int(t["id"]),"department":t["department_name"],"output":t["output_name"],"plan":float(t["target_quantity"] or 0),"actual":float(t["actual_quantity"] or 0),"input":c["name"],"required":required,"available":available,"open_replenishment":open_qty,"shortage":max(0.0,required-available-open_qty),"unit":c["default_unit"] or "шт","due_at":t.get("due_at") or ""})
    return rows

def create_xlsx(chat_id:int,user_id:int,*,start_date:str|None=None,end_date:str|None=None)->Path:
    rows=build_rows(chat_id,user_id,start_date=start_date,end_date=end_date);path=reporting.reports_dir()/f"plan_need_stock_{datetime.now():%Y%m%d_%H%M%S}.xlsx";wb=Workbook();ws=wb.active;ws.title="План и потребность";headers=["Задание","Отдел","Выпуск","План","Факт","Потребность","Нужно","Доступно","В пополнении","Дефицит","Ед.","Срок"]
    ws.append(headers)
    for c in ws[1]:c.font=Font(bold=True);c.alignment=Alignment(wrap_text=True)
    for r in rows:ws.append([r["task_id"],r["department"],r["output"],r["plan"],r["actual"],r["input"],r["required"],r["available"],r["open_replenishment"],r["shortage"],r["unit"],r["due_at"]])
    for col in range(1,len(headers)+1):ws.column_dimensions[get_column_letter(col)].width=min(32,max(10,max((len(str(ws.cell(row=i,column=col).value or "")) for i in range(1,ws.max_row+1)),default=10)+2))
    ws.freeze_panes="A2";wb.save(path);return path

def create_pdf(chat_id:int,user_id:int,*,start_date:str|None=None,end_date:str|None=None)->Path:
    rows=build_rows(chat_id,user_id,start_date=start_date,end_date=end_date);path=reporting.reports_dir()/f"plan_need_stock_{datetime.now():%Y%m%d_%H%M%S}.pdf";font=reporting._register_pdf_font();doc=SimpleDocTemplate(str(path),pagesize=landscape(A4),leftMargin=12,rightMargin=12,topMargin=14,bottomMargin=14);styles=getSampleStyleSheet();title=ParagraphStyle("ru",parent=styles["Heading2"],fontName=font,fontSize=12)
    data=[["№","Отдел","Выпуск","План","Факт","Потребность","Нужно","Доступно","Пополнение","Дефицит"]]
    for r in rows:data.append([str(r["task_id"]),str(r["department"]),str(r["output"]),f'{r["plan"]:g}',f'{r["actual"]:g}',str(r["input"]),f'{r["required"]:g}',f'{r["available"]:g}',f'{r["open_replenishment"]:g}',f'{r["shortage"]:g}'])
    table=Table(data,repeatRows=1);table.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),font),("FONTSIZE",(0,0),(-1,-1),7),("BACKGROUND",(0,0),(-1,0),colors.lightgrey),("GRID",(0,0),(-1,-1),0.25,colors.grey),("VALIGN",(0,0),(-1,-1),"TOP")]))
    doc.build([Paragraph("План → потребность → наличие → дефицит → факт",title),table]);return path
