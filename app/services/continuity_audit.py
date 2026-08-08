from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from .. import db
from . import repository as repo


def _sheet(ws, rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> None:
    ws.append([title for _, title in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(key, "") for key, _ in columns])
    for index, (_, title) in enumerate(columns, start=1):
        ws.column_dimensions[chr(64 + index) if index <= 26 else 'A'].width = min(max(len(title) + 2, 12), 34)
    ws.freeze_panes = "A2"


def build_continuity_audit_xlsx(chat_id: int) -> bytes:
    scope = repo.resolve_scope_chat_id(chat_id)
    packages = [dict(r) for r in db.fetchall(
        """SELECT p.*,a.name AS area_name FROM shift_sync_packages p
        LEFT JOIN areas a ON a.id=p.area_id WHERE p.chat_id=? ORDER BY p.id DESC""", (scope,)
    )]
    package_ids = [int(x["id"]) for x in packages]
    if package_ids:
        marks = ",".join("?" for _ in package_ids)
        items = [dict(r) for r in db.fetchall(
            f"SELECT * FROM shift_sync_items WHERE package_id IN ({marks}) ORDER BY package_id,sequence_no,id", package_ids
        )]
    else:
        items = []
    handovers = [dict(r) for r in db.fetchall(
        """SELECT h.*,a.name AS area_name FROM shift_handovers h
        LEFT JOIN areas a ON a.id=h.area_id WHERE h.chat_id=? ORDER BY h.id DESC""", (scope,)
    )]
    checks = []
    handover_ids = [int(x["id"]) for x in handovers]
    if handover_ids:
        marks = ",".join("?" for _ in handover_ids)
        checks = [dict(r) for r in db.fetchall(
            f"SELECT * FROM shift_handover_checks WHERE handover_id IN ({marks}) ORDER BY handover_id,sort_order,id", handover_ids
        )]
    devices = [dict(r) for r in db.fetchall(
        "SELECT * FROM miniapp_devices WHERE last_chat_id=? ORDER BY last_seen_at DESC", (scope,)
    )]
    reminders = [dict(r) for r in db.fetchall(
        "SELECT * FROM shift_continuity_reminders WHERE chat_id=? ORDER BY id DESC", (scope,)
    )]

    wb = Workbook()
    ws = wb.active
    ws.title = "Пакеты"
    _sheet(ws, packages, [
        ("id","ID"),("user_id","Сотрудник ID"),("area_name","Площадка"),("status","Статус"),
        ("item_count","Записей"),("accepted_count","Принято"),("review_count","На проверке"),
        ("rejected_count","Отклонено"),("error_count","Ошибки"),("submitted_at","Отправлено"),("reviewed_at","Проверено"),
    ])
    ws = wb.create_sheet("Записи пакетов")
    _sheet(ws, items, [
        ("id","ID"),("package_id","Пакет"),("sequence_no","№"),("status","Статус"),("operation_id","Операция"),
        ("message","Сообщение"),("client_request_id","Защита от дубля"),("created_at","Создано"),("updated_at","Изменено"),
    ])
    ws = wb.create_sheet("Передачи смен")
    _sheet(ws, handovers, [
        ("id","ID"),("from_user_id","От кого"),("to_user_id","Кому"),("area_name","Площадка"),("status","Статус"),
        ("summary","Комментарий"),("unfinished_count","Незавершено"),("issue_count","Проблем"),("created_at","Создано"),("acknowledged_at","Принято"),
    ])
    ws = wb.create_sheet("Чек-листы")
    _sheet(ws, checks, [
        ("handover_id","Передача"),("label","Пункт"),("is_required","Обязательный"),("is_checked","Отмечен"),
        ("note","Замечание"),("checked_by","Кем отмечен"),("checked_at","Когда"),
    ])
    ws = wb.create_sheet("Устройства")
    _sheet(ws, devices, [
        ("user_id","Пользователь"),("device_id","Устройство ID"),("device_name","Название"),("platform","Платформа"),
        ("first_seen_at","Первый вход"),("last_seen_at","Последний вход"),("revoked_at","Отозвано"),("revoke_reason","Причина"),
    ])
    ws = wb.create_sheet("Напоминания")
    _sheet(ws, reminders, [
        ("reminder_kind","Тип"),("related_id","Объект"),("recipient_user_id","Получатель"),("reminder_level","Уровень"),("created_at","Создано"),
    ])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
