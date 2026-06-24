from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable
import html
import os
import re
import shutil
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from .. import db
from ..config import settings
from .normalize import normalize_key


OPERATION_LABELS: dict[str, str] = {
    "production": "Производство",
    "material_in": "Поступление сырья",
    "material_out": "Расход сырья",
    "energy": "Электроэнергия",
    "assembly": "Сборка",
    "shipment": "Отгрузка",
    "stock_in": "Склад: приход",
    "stock_out": "Склад: уход",
    "inventory_adjust": "Инвентаризация",
}

ENTITY_LABELS: dict[str, str] = {
    "component": "Комплектующие",
    "product": "Готовая продукция",
    "material": "Сырьё",
    "meter": "Прибор учёта",
    "stock_item": "Складская позиция",
}


def _fmt_number(value: object, decimals: int = 3) -> str:
    """Показывает числа обычным видом: 4 000 000, а не 4e+06."""
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return ""
        try:
            decimal_value = Decimal(stripped.replace(" ", "").replace(",", "."))
        except (InvalidOperation, ValueError):
            return value
    else:
        try:
            decimal_value = Decimal(str(value))
        except (InvalidOperation, ValueError):
            return str(value)
    if not decimal_value.is_finite():
        return str(value)
    quant = Decimal(1).scaleb(-decimals)
    decimal_value = decimal_value.quantize(quant).normalize()
    sign = "-" if decimal_value < 0 else ""
    decimal_value = abs(decimal_value)
    plain = format(decimal_value, "f")
    if "." in plain:
        integer, fraction = plain.split(".", 1)
        fraction = fraction.rstrip("0")
    else:
        integer, fraction = plain, ""
    grouped = f"{int(integer or '0'):,}".replace(",", " ")
    return sign + grouped + (("," + fraction) if fraction else "")


def _display_cell(value: object) -> str:
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return _fmt_number(value)
    return "" if value is None else str(value)


@dataclass(frozen=True)
class PeriodFilter:
    title: str
    where_sql: str
    params: tuple[str, ...]
    error: str = ""


_DATE_PATTERN = re.compile(r"(?<!\d)(\d{1,2})[./](\d{1,2})[./](\d{2,4})(?!\d)")
_MAX_PERIOD_DAYS = 36525


def _format_ru_day(value: date) -> str:
    return value.strftime("%d.%m.%Y")


def _parse_ru_day(match: re.Match[str]) -> date | None:
    day = int(match.group(1))
    month = int(match.group(2))
    year_raw = match.group(3)
    year = int(year_raw)
    if len(year_raw) == 2:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _period_from_days(days: list[date]) -> PeriodFilter | None:
    if not days:
        return None
    start = days[0]
    end = days[1] if len(days) > 1 else days[0]
    if start > end:
        return PeriodFilter(
            "",
            "1=0",
            tuple(),
            "Дата начала не может быть позже даты конца. Укажите период так: 12.05.2022-20.06.2022.",
        )
    if (end - start).days > _MAX_PERIOD_DAYS:
        return PeriodFilter("", "1=0", tuple(), "Период не должен быть длиннее 100 лет.")
    end_exclusive = end + timedelta(days=1)
    if start == end:
        title = f"за {_format_ru_day(start)}"
    else:
        title = f"с {_format_ru_day(start)} по {_format_ru_day(end)}"
    return PeriodFilter(title, "o.created_at >= ? AND o.created_at < ?", (str(start), str(end_exclusive)))


def period_error_for_text(text: str) -> str:
    return _date_bounds_for_text(text).error


def looks_like_period_text(text: str) -> bool:
    return bool(_DATE_PATTERN.search(text or ""))


def _date_bounds_for_text(text: str) -> PeriodFilter:
    parsed_days: list[date] = []
    invalid_date_seen = False
    for match in _DATE_PATTERN.finditer(text or ""):
        parsed = _parse_ru_day(match)
        if parsed is None:
            invalid_date_seen = True
        else:
            parsed_days.append(parsed)
    if invalid_date_seen and not parsed_days:
        return PeriodFilter("", "1=0", tuple(), "Дата указана неверно. Используйте формат 12.05.2022.")
    explicit = _period_from_days(parsed_days[:2])
    if explicit is not None:
        return explicit

    key = normalize_key(text)
    today = datetime.now().date()
    if "вчера" in key:
        start = today - timedelta(days=1)
        end = today
        return PeriodFilter("за вчера", "o.created_at >= ? AND o.created_at < ?", (str(start), str(end)))
    if "недел" in key or "7" in key and "дн" in key:
        start = today - timedelta(days=7)
        return PeriodFilter("за 7 дней", "o.created_at >= ?", (str(start),))
    if "месяц" in key or "мес" in key:
        start = today.replace(day=1)
        return PeriodFilter("за текущий месяц", "o.created_at >= ?", (str(start),))
    if "сегодня" in key or not any(word in key for word in ("вчера", "недел", "месяц", "все", "всё", "общ")):
        return PeriodFilter("за сегодня", "o.created_at >= ?", (str(today),))
    return PeriodFilter("за всё время", "1=1", tuple())


def reports_dir() -> Path:
    path = settings.data_dir / "reports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _safe_name(value: str) -> str:
    clean = re.sub(r"[^0-9A-Za-zА-Яа-яёЁ_.-]+", "_", value).strip("_")
    return clean[:80] or "report"


def inventory_rows(chat_id: int) -> list[dict]:
    rows = db.fetchall(
        """
        SELECT i.quantity,i.unit,i.entity_type,e.name AS entity_name,a.name AS area_name
        FROM inventory i
        LEFT JOIN entities e ON e.id=i.entity_id
        LEFT JOIN areas a ON a.id=i.area_id
        WHERE i.chat_id=? AND ABS(i.quantity) > 0.000001
        ORDER BY i.entity_type, COALESCE(a.name,''), e.name, i.unit
        """,
        (chat_id,),
    )
    return [dict(r) for r in rows]


def operation_rows(chat_id: int, period: PeriodFilter) -> list[dict]:
    rows = db.fetchall(
        f"""
        SELECT o.operation_type,o.entity_type,e.name AS entity_name,a.name AS area_name,o.unit,
               SUM(o.quantity) AS total_quantity, COUNT(o.id) AS count_rows
        FROM operations o
        LEFT JOIN entities e ON e.id=o.entity_id
        LEFT JOIN areas a ON a.id=o.area_id
        WHERE o.chat_id=? AND {period.where_sql}
        GROUP BY o.operation_type,o.entity_type,o.entity_id,o.area_id,o.unit
        ORDER BY o.operation_type, COALESCE(a.name,''), e.name
        """,
        (chat_id, *period.params),
    )
    return [dict(r) for r in rows]


def raw_operation_rows(chat_id: int, period: PeriodFilter, limit: int = 5000) -> list[dict]:
    rows = db.fetchall(
        f"""
        SELECT o.created_at,o.operation_type,o.entity_type,e.name AS entity_name,a.name AS area_name,
               o.quantity,o.unit,o.user_id,o.group_chat_id,o.raw_text
        FROM operations o
        LEFT JOIN entities e ON e.id=o.entity_id
        LEFT JOIN areas a ON a.id=o.area_id
        WHERE o.chat_id=? AND {period.where_sql}
        ORDER BY o.created_at DESC
        LIMIT ?
        """,
        (chat_id, *period.params, limit),
    )
    return [dict(r) for r in rows]



def build_stock_text(chat_id: int) -> str:
    rows = inventory_rows(chat_id)
    if not rows:
        return "Склад пока пустой."
    lines = ["Склад"]
    current = None
    for row in rows[:80]:
        title = ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or "Позиции")
        if title != current:
            current = title
            lines.append(f"\n{title}:")
        name = row.get("entity_name") or "Позиция"
        area = f" · {row['area_name']}" if row.get("area_name") else ""
        qty = float(row.get("quantity") or 0)
        lines.append(f"• {name} — {_fmt_number(qty)} {row.get('unit') or ''}{area}")
    if len(rows) > 80:
        lines.append(f"\nЕщё строк: {len(rows) - 80}")
    return "\n".join(lines)

def build_text_report(chat_id: int, request_text: str, user_id: int | None = None) -> str:
    from . import repository as repo

    period = _date_bounds_for_text(request_text)
    if period.error:
        return period.error
    prefs = repo.get_export_preferences(chat_id, user_id)
    lines: list[str] = [f"Отчёт {period.title}"]

    if prefs.get("period_totals"):
        ops = operation_rows(chat_id, period)
        if not ops:
            lines.append("\nДвижения за период пока нет.")
        else:
            current = None
            for row in ops:
                title = OPERATION_LABELS.get(row.get("operation_type") or "", "Запись")
                if title != current:
                    current = title
                    lines.append(f"\n{title}:")
                name = row.get("entity_name") or ENTITY_LABELS.get(row.get("entity_type") or "", "Позиция")
                area = f" · {row['area_name']}" if row.get("area_name") else ""
                qty = float(row.get("total_quantity") or 0)
                lines.append(f"• {name} — {_fmt_number(qty)} {row.get('unit') or ''}{area}")

    if prefs.get("daily_matrix"):
        labels, matrix = movement_matrix_rows(chat_id, period)
        if labels and matrix:
            lines.append("\nПо датам:")
            shown = labels[-10:]
            for label in shown:
                total = sum(float(row["values"].get(label) or 0) for row in matrix)
                lines.append(f"• {label}: {_fmt_number(total)}")
            if len(labels) > len(shown):
                lines.append("• Полная таблица доступна в скачанном отчёте.")

    if prefs.get("capacity"):
        lines.append("\n" + build_all_assembly_capacity_report(chat_id))

    if prefs.get("inventory"):
        inv = inventory_rows(chat_id)
        if inv:
            lines.append("\nОстатки склада:")
            for row in inv[:25]:
                name = row.get("entity_name") or ENTITY_LABELS.get(row.get("entity_type") or "", "Позиция")
                area = f" · {row['area_name']}" if row.get("area_name") else ""
                qty = float(row.get("quantity") or 0)
                lines.append(f"• {name} — {_fmt_number(qty)} {row.get('unit') or ''}{area}")
            if len(inv) > 25:
                lines.append(f"• Ещё строк: {len(inv) - 25}")
        else:
            lines.append("\nСклад пока пустой.")

    if prefs.get("journal"):
        journal = raw_operation_rows(chat_id, period, limit=10)
        if journal:
            lines.append("\nПоследние записи:")
            for row in journal:
                title = OPERATION_LABELS.get(row.get("operation_type") or "", "Запись")
                name = row.get("entity_name") or ENTITY_LABELS.get(row.get("entity_type") or "", "Позиция")
                qty = float(row.get("quantity") or 0)
                lines.append(f"• {row.get('created_at') or ''} · {title}: {name} — {_fmt_number(qty)} {row.get('unit') or ''}")

    if len(lines) == 1:
        lines.append("\nВыберите хотя бы один раздел.")
    return "\n".join(lines)


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for column_cells in ws.columns:
        col = get_column_letter(column_cells[0].column)
        max_len = 10
        for cell in column_cells:
            max_len = max(max_len, min(len(str(cell.value or "")), 45))
        ws.column_dimensions[col].width = max_len + 2
    ws.freeze_panes = "A2"


def create_xlsx_report(chat_id: int, request_text: str = "отчёт") -> Path:
    period = _date_bounds_for_text(request_text)
    if period.error:
        raise ValueError(period.error)
    wb = Workbook()
    ws = wb.active
    ws.title = "Склад"
    ws.append(["Тип", "Название", "Участок", "Количество", "Ед."])
    for row in inventory_rows(chat_id):
        ws.append([
            ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or ""),
            row.get("entity_name") or "",
            row.get("area_name") or "Общий склад",
            float(row.get("quantity") or 0),
            row.get("unit") or "",
        ])
    _style_sheet(ws)

    ws2 = wb.create_sheet("Итоги за период")
    ws2.append(["Период", period.title])
    ws2.append([])
    ws2.append(["Операция", "Тип", "Название", "Участок", "Количество", "Ед.", "Строк"])
    for row in operation_rows(chat_id, period):
        ws2.append([
            OPERATION_LABELS.get(row.get("operation_type") or "", row.get("operation_type") or ""),
            ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or ""),
            row.get("entity_name") or "",
            row.get("area_name") or "",
            float(row.get("total_quantity") or 0),
            row.get("unit") or "",
            int(row.get("count_rows") or 0),
        ])
    _style_sheet(ws2)

    ws3 = wb.create_sheet("Журнал")
    ws3.append(["Дата", "Операция", "Тип", "Название", "Участок", "Количество", "Ед.", "Работник", "Группа"])
    for row in raw_operation_rows(chat_id, period):
        ws3.append([
            row.get("created_at") or "",
            OPERATION_LABELS.get(row.get("operation_type") or "", row.get("operation_type") or ""),
            ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or ""),
            row.get("entity_name") or "",
            row.get("area_name") or "",
            float(row.get("quantity") or 0),
            row.get("unit") or "",
            row.get("user_id") or "",
            row.get("group_chat_id") or "",
        ])
    _style_sheet(ws3)

    filename = f"uchet_{_safe_name(period.title)}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    path = reports_dir() / filename
    wb.save(path)
    return path



def _font_has_cyrillic(font_path: Path) -> bool:
    try:
        probe = TTFont("ProbeFont", str(font_path))
        glyphs = getattr(probe.face, "charToGlyph", {})
        return all(ord(ch) in glyphs for ch in "Производственный отчёт Склад")
    except Exception:
        return False


def _add_font_path(candidates: list[Path], value: str | Path | None) -> None:
    if value is None:
        return
    raw = str(value).strip().strip('"')
    if not raw:
        return
    path = Path(raw).expanduser()
    if path.is_dir():
        wanted = ("dejavu", "noto", "free", "liberation", "arial", "segoe", "calibri", "sans")
        for item in path.rglob("*"):
            if item.suffix.lower() not in {".ttf", ".ttc", ".otf"}:
                continue
            if any(part in item.name.lower() for part in wanted):
                candidates.append(item)
        return
    candidates.append(path)


def _font_paths_from_fontconfig() -> list[Path]:
    if not shutil.which("fc-match"):
        return []
    result: list[Path] = []
    families = ("DejaVu Sans", "Noto Sans", "Liberation Sans", "Arial", "FreeSans", "sans")
    for family in families:
        try:
            completed = subprocess.run(
                ["fc-match", "-f", "%{file}", family],
                check=False,
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                text=True,
                timeout=3,
            )
        except Exception:
            continue
        value = (completed.stdout or "").strip()
        if value:
            result.append(Path(value))
    return result


def _pdf_font_candidates() -> list[Path]:
    candidates: list[Path] = []
    for env_name in ("REPORT_PDF_FONT", "REPORT_PDF_FONT_DIR"):
        env_value = os.environ.get(env_name, "").strip()
        if env_value:
            for part in env_value.split(os.pathsep):
                _add_font_path(candidates, part)

    project_root = Path(__file__).resolve().parents[2]
    direct_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/usr/local/share/fonts/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
        "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
        "/app/.fonts/DejaVuSans.ttf",
        "/app/fonts/DejaVuSans.ttf",
        "/opt/render/project/src/fonts/DejaVuSans.ttf",
        "/workspace/fonts/DejaVuSans.ttf",
        "/var/task/fonts/DejaVuSans.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/Arial.ttf",
        "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/Library/Fonts/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ]
    for item in direct_paths:
        _add_font_path(candidates, item)

    search_roots = [
        project_root / "fonts",
        project_root / ".fonts",
        settings.data_dir / "fonts",
        Path("/usr/share/fonts"),
        Path("/usr/local/share/fonts"),
        Path("/app/fonts"),
        Path("/app/.fonts"),
        Path("/opt/render/project/src/fonts"),
        Path("/workspace/fonts"),
        Path("/var/task/fonts"),
        Path.home() / ".fonts",
        Path.home() / ".local/share/fonts",
    ]
    for root in search_roots:
        _add_font_path(candidates, root)

    candidates.extend(_font_paths_from_fontconfig())

    result: list[Path] = []
    seen: set[str] = set()
    for item in candidates:
        key = str(item.resolve() if item.exists() else item)
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _register_pdf_font() -> str:
    for font_path in _pdf_font_candidates():
        if not font_path.exists() or not _font_has_cyrillic(font_path):
            continue
        try:
            font_name = "BotSans" + str(abs(hash(str(font_path))) % 100000)
            if font_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
            return font_name
        except Exception:
            continue
    raise ValueError("PDF не собран: на сервере нет русского шрифта. Запустите start_linux.sh заново или установите fonts-dejavu-core. Ещё можно положить DejaVuSans.ttf или NotoSans-Regular.ttf в папку fonts рядом с ботом.")


def pdf_font_status() -> tuple[bool, str]:
    try:
        font_name = _register_pdf_font()
        return True, f"PDF готов: русский шрифт найден ({font_name})."
    except ValueError as exc:
        return False, str(exc)


def create_pdf_report(chat_id: int, request_text: str = "отчёт") -> Path:
    period = _date_bounds_for_text(request_text)
    if period.error:
        raise ValueError(period.error)
    font_name = _register_pdf_font()
    filename = f"uchet_{_safe_name(period.title)}_{datetime.now():%Y%m%d_%H%M%S}.pdf"
    path = reports_dir() / filename
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleRu", parent=styles["Title"], fontName=font_name, fontSize=16, leading=20)
    normal_style = ParagraphStyle("NormalRu", parent=styles["Normal"], fontName=font_name, fontSize=9, leading=12)
    story = [Paragraph(f"Производственный отчёт {period.title}", title_style), Spacer(1, 6)]

    inv_data = [["Тип", "Название", "Участок", "Количество", "Ед."]]
    for row in inventory_rows(chat_id):
        inv_data.append([
            ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or ""),
            row.get("entity_name") or "",
            row.get("area_name") or "Общий склад",
            _fmt_number(float(row.get('quantity') or 0)),
            row.get("unit") or "",
        ])
    if len(inv_data) == 1:
        inv_data.append(["", "Склад пока пустой", "", "", ""])
    story.append(Paragraph("Склад", normal_style))
    story.append(_pdf_table(inv_data, font_name))
    story.append(Spacer(1, 8))

    ops_data = [["Операция", "Название", "Участок", "Количество", "Ед.", "Строк"]]
    for row in operation_rows(chat_id, period):
        ops_data.append([
            OPERATION_LABELS.get(row.get("operation_type") or "", row.get("operation_type") or ""),
            row.get("entity_name") or ENTITY_LABELS.get(row.get("entity_type") or "", ""),
            row.get("area_name") or "",
            _fmt_number(float(row.get('total_quantity') or 0)),
            row.get("unit") or "",
            str(row.get("count_rows") or 0),
        ])
    if len(ops_data) == 1:
        ops_data.append(["", "Движения за период пока нет", "", "", "", ""])
    story.append(Paragraph("Итоги за период", normal_style))
    story.append(_pdf_table(ops_data, font_name))
    doc.build(story)
    return path


def _pdf_table(data: list[list[str]], font_name: str) -> Table:
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
    ]))
    return table


def build_assembly_capacity_report(chat_id: int, request_text: str) -> str:
    from .matcher import confident_match
    from . import repository as repo

    # Убираем служебные слова, чтобы остался запрос по изделию.
    key_text = request_text
    for word in ["сколько", "можно", "собрать", "изготовить", "сделать", "готово", "изделий", "изделие"]:
        key_text = re.sub(rf"\b{word}\b", " ", key_text, flags=re.IGNORECASE)
    match, _ = confident_match(chat_id, key_text.strip() or request_text, allowed_types={"product"})
    if not match:
        return "Не нашёл изделие. Уточните название."
    components = repo.list_product_components(match.target_id)
    if not components:
        return f"У изделия «{match.name}» пока не задан состав."
    lines = [f"Можно собрать: {match.name}"]
    possible_values: list[float] = []
    missing: list[str] = []
    for comp in components:
        row = db.fetchone(
            """
            SELECT COALESCE(SUM(quantity),0) AS qty
            FROM inventory
            WHERE chat_id=? AND entity_type='component' AND entity_id=?
            """,
            (chat_id, int(comp["component_id"])),
        )
        stock_qty = float(row["qty"] if row else 0)
        need = float(comp["quantity"] or 0)
        can_make = stock_qty // need if need > 0 else 0
        possible_values.append(can_make)
        lines.append(f"• {comp['name']}: есть {_fmt_number(stock_qty)}, нужно {_fmt_number(need)} на 1")
        if stock_qty < need:
            missing.append(f"• {comp['name']} — не хватает {_fmt_number(need - stock_qty)}")
    possible = int(min(possible_values)) if possible_values else 0
    lines.insert(1, f"Итого сейчас: {_fmt_number(possible)} шт.")
    if missing:
        lines.append("\nДля одной полной сборки не хватает:")
        lines.extend(missing)
    return "\n".join(lines)

# --- Дополнение: общие комплектующие, расчёт сборки по всем изделиям и настраиваемые файлы ---


def _period_start_end(period: PeriodFilter) -> tuple[date, date]:
    today = datetime.now().date()
    if len(period.params) >= 2:
        start = date.fromisoformat(str(period.params[0])[:10])
        exclusive_end = date.fromisoformat(str(period.params[1])[:10])
        return start, exclusive_end - timedelta(days=1)
    if len(period.params) == 1:
        start = date.fromisoformat(str(period.params[0])[:10])
        return start, today
    return today, today


def _bucket_labels_for_period(period: PeriodFilter) -> tuple[str, list[str]]:
    start, end = _period_start_end(period)
    day_count = (end - start).days + 1
    if day_count <= 62:
        labels: list[str] = []
        current = start
        while current <= end:
            labels.append(current.strftime("%d.%m.%Y"))
            current += timedelta(days=1)
        return "day", labels
    labels = []
    current = date(start.year, start.month, 1)
    last = date(end.year, end.month, 1)
    while current <= last:
        labels.append(current.strftime("%m.%Y"))
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    return "month", labels


def movement_matrix_rows(chat_id: int, period: PeriodFilter) -> tuple[list[str], list[dict]]:
    bucket_type, labels = _bucket_labels_for_period(period)
    if bucket_type == "day":
        bucket_sql = "strftime('%d.%m.%Y', o.created_at)"
    else:
        bucket_sql = "strftime('%m.%Y', o.created_at)"
    rows = db.fetchall(
        f"""
        SELECT o.operation_type,o.entity_type,e.name AS entity_name,o.unit,{bucket_sql} AS bucket,
               SUM(o.quantity) AS total_quantity
        FROM operations o
        LEFT JOIN entities e ON e.id=o.entity_id
        WHERE o.chat_id=? AND {period.where_sql}
        GROUP BY o.operation_type,o.entity_type,o.entity_id,o.unit,bucket
        ORDER BY o.operation_type,e.name,bucket
        """,
        (chat_id, *period.params),
    )
    grouped: dict[tuple[str, str, str, str], dict] = {}
    for raw_row in rows:
        row = dict(raw_row)
        key = (
            str(row.get("operation_type") or ""),
            str(row.get("entity_type") or ""),
            str(row.get("entity_name") or ""),
            str(row.get("unit") or ""),
        )
        item = grouped.setdefault(key, {
            "operation_type": key[0],
            "entity_type": key[1],
            "entity_name": key[2],
            "unit": key[3],
            "values": {label: 0.0 for label in labels},
        })
        bucket = str(row.get("bucket") or "")
        if bucket in item["values"]:
            item["values"][bucket] = float(row.get("total_quantity") or 0)
    return labels, list(grouped.values())


def _component_stock(chat_id: int, component_id: int) -> float:
    row = db.fetchone(
        """
        SELECT COALESCE(SUM(quantity),0) AS qty
        FROM inventory
        WHERE chat_id=? AND entity_type='component' AND entity_id=?
        """,
        (chat_id, component_id),
    )
    return float(row["qty"] if row else 0)


def product_capacity_rows(chat_id: int) -> list[dict]:
    from . import repository as repo

    rows: list[dict] = []
    for product_pack in repo.all_products_with_components(chat_id):
        product = product_pack["product"]
        components = product_pack["components"]
        if not components:
            rows.append({
                "product_name": product.name,
                "possible": "состав не задан",
                "component_name": "",
                "stock": "",
                "need": "",
                "missing_next": "",
                "unit": "",
            })
            continue
        possible_values: list[float] = []
        component_info: list[dict] = []
        for comp in components:
            stock_qty = _component_stock(chat_id, int(comp["component_id"]))
            need = float(comp["quantity"] or 0)
            can_make = stock_qty // need if need > 0 else 0
            possible_values.append(can_make)
            component_info.append({
                "name": str(comp.get("name") or ""),
                "stock": stock_qty,
                "need": need,
                "unit": str(comp.get("default_unit") or "шт"),
            })
        possible = int(min(possible_values)) if possible_values else 0
        for comp in component_info:
            missing_next = max(0.0, comp["need"] * (possible + 1) - comp["stock"])
            rows.append({
                "product_name": product.name,
                "possible": possible,
                "component_name": comp["name"],
                "stock": comp["stock"],
                "need": comp["need"],
                "missing_next": missing_next,
                "unit": comp["unit"],
            })
    return rows


def _format_target_rows_text(rows: list[list[object]], title: str = "План по количеству") -> str:
    if not rows:
        return "План пустой. Укажите изделие и нужное количество."
    lines = [title]
    for product, target, possible, component, stock, need_one, need_total, missing, unit in rows[:80]:
        if component and component != "состав не задан":
            lines.append(
                f"• {product} · {_fmt_number(target)} шт: {component} — нужно {_fmt_number(need_total)}, "
                f"есть {_fmt_number(stock)}, не хватает {_fmt_number(missing)} {unit}"
            )
        else:
            lines.append(f"• {product} · {_fmt_number(target)} шт: состав пока не задан.")
    if len(rows) > 80:
        lines.append("• Полный план есть в скачанном отчёте.")
    return "\n".join(lines)


def build_all_assembly_capacity_report(chat_id: int, targets: list[int] | None = None) -> str:
    rows = product_capacity_rows(chat_id)
    if not rows:
        return "Изделия пока не созданы."
    lines = ["Расчёт сборки по изделиям"]
    current = None
    for row in rows:
        product_name = str(row["product_name"])
        if product_name != current:
            current = product_name
            lines.append(f"\n{product_name}: можно собрать {_fmt_number(row['possible'])}")
        if row.get("component_name"):
            missing = float(row.get("missing_next") or 0)
            extra = f", для ещё 1 не хватает {_fmt_number(missing)} {row.get('unit') or ''}" if missing > 0 else ""
            lines.append(f"• {row['component_name']}: есть {_fmt_number(row['stock'])}, нужно {_fmt_number(row['need'])}{extra}")
        else:
            lines.append("• Состав пока не задан.")
    target_rows = _product_target_rows(chat_id, targets or DEFAULT_ASSEMBLY_TARGETS)
    if target_rows:
        lines.append("\n" + _format_target_rows_text(target_rows))
    return "\n".join(lines)


def build_assembly_capacity_report(chat_id: int, request_text: str) -> str:
    from .matcher import confident_match
    from . import repository as repo

    targets = _target_quantities_from_text(request_text, include_defaults=False)
    key = normalize_key(request_text)
    if any(word in key for word in ("все", "всё", "кажд", "общ")) or key.strip() in {"сколько можно собрать", "расчет сборки", "расчёт сборки"}:
        return build_all_assembly_capacity_report(chat_id, targets or DEFAULT_ASSEMBLY_TARGETS)
    key_text = _strip_target_numbers(request_text)
    for word in ["сколько", "можно", "собрать", "сбора", "сбор", "сборки", "изготовить", "сделать", "готово", "изделий", "изделие", "расчет", "расчёт", "нужно", "не", "хватает", "до", "для", "план", "цель", "цели"]:
        key_text = re.sub(rf"\b{word}\b", " ", key_text, flags=re.IGNORECASE)
    match, _ = confident_match(chat_id, key_text.strip() or request_text, allowed_types={"product"})
    if not match:
        custom_rows = _product_target_rows_for_request(chat_id, request_text, include_defaults=False)
        if custom_rows:
            return _format_target_rows_text(custom_rows)
        return build_all_assembly_capacity_report(chat_id, targets or DEFAULT_ASSEMBLY_TARGETS)
    components = repo.list_product_components(match.target_id)
    if not components:
        return f"У изделия «{match.name}» пока не задан состав."
    possible_values: list[float] = []
    lines = [f"Можно собрать: {match.name}"]
    for comp in components:
        stock_qty = _component_stock(chat_id, int(comp["component_id"]))
        need = float(comp["quantity"] or 0)
        can_make = stock_qty // need if need > 0 else 0
        possible_values.append(can_make)
        lines.append(f"• {comp['name']}: есть {_fmt_number(stock_qty)}, нужно {_fmt_number(need)} на 1")
    possible = int(min(possible_values)) if possible_values else 0
    lines.insert(1, f"Итого сейчас: {_fmt_number(possible)} шт.")
    missing_next: list[str] = []
    for comp in components:
        stock_qty = _component_stock(chat_id, int(comp["component_id"]))
        need = float(comp["quantity"] or 0)
        missing = max(0.0, need * (possible + 1) - stock_qty)
        if missing > 0:
            missing_next.append(f"• {comp['name']} — {_fmt_number(missing)} {comp.get('default_unit') or 'шт'}")
    if missing_next:
        lines.append("\nДля ещё 1 изделия не хватает:")
        lines.extend(missing_next)
    target_rows = _product_target_rows_for_request(chat_id, request_text, include_defaults=True)
    own_rows = [row for row in target_rows if str(row[0]) == str(match.name)]
    if own_rows:
        lines.append("\nДо нужного количества:")
        for _product, target, _possible, component, stock, _need_one, need_total, missing, unit in own_rows[:45]:
            lines.append(f"• {_fmt_number(target)} шт · {component}: нужно {_fmt_number(need_total)}, есть {_fmt_number(stock)}, не хватает {_fmt_number(missing)} {unit}")
    return "\n".join(lines)



def _number_format_for_header(header: str) -> str:
    if header in {"Количество", "Есть", "Нужно на 1", "Не хватает для ещё 1", "Итого", "Цель", "Можно собрать", "Нужно на цель", "Не хватает", "Собрано", "Отправлено/продано", "Остаток"}:
        return "#,##0.###"
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", header or "") or re.fullmatch(r"\d{2}\.\d{4}", header or ""):
        return "#,##0.###"
    if header == "Строк":
        return "0"
    return "General"


def _style_used_range(ws, header_rows: set[int] | None = None, section_rows: set[int] | None = None) -> None:
    header_rows = header_rows or {1}
    section_rows = section_rows or set()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="EEF6FC")
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        row_number = row[0].row
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_number == 1:
                cell.font = title_font
            elif row_number in section_rows:
                cell.fill = section_fill
                cell.font = section_font
            elif row_number in header_rows:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            elif any(c.value not in (None, "") for c in row):
                cell.border = border
    for column_cells in ws.columns:
        col = get_column_letter(column_cells[0].column)
        max_len = 10
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)), 38))
        ws.column_dimensions[col].width = min(max_len + 2, 42)
    for row_index in range(1, max_row + 1):
        ws.row_dimensions[row_index].height = 20 if row_index not in section_rows else 24
    ws.freeze_panes = "A3" if max_row >= 3 else "A2"


def _append_report_section(ws, title: str, header: list[str], rows: list[list[object]], start_row: int) -> tuple[int, int, int]:
    section_row = start_row
    ws.cell(row=section_row, column=1, value=title)
    header_row = section_row + 1
    for col_index, value in enumerate(header, start=1):
        ws.cell(row=header_row, column=col_index, value=value)
    body_rows = rows or [_empty_row(len(header))]
    for offset, row in enumerate(body_rows, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=header_row + offset, column=col_index, value=value)
            if col_index <= len(header):
                cell.number_format = _number_format_for_header(str(header[col_index - 1]))
    return section_row, header_row, header_row + len(body_rows) + 2


def _full_export_selected() -> dict[str, bool]:
    return {"inventory": True, "period_totals": True, "daily_matrix": True, "capacity": True, "journal": True}


def create_xlsx_report(chat_id: int, request_text: str = "отчёт", user_id: int | None = None) -> Path:
    period = _date_bounds_for_text(request_text)
    if period.error:
        raise ValueError(period.error)
    wb = Workbook()
    sections = report_sections(chat_id, request_text, user_id=user_id, selected=_full_export_selected())
    if not sections:
        sections = [("Отчёт", ["Раздел", "Состояние"], [["Отчёт", "Нет данных"]])]

    summary = wb.active
    summary.title = "Отчёт"
    max_width = max((len(header) for _, header, _ in sections), default=2)
    summary.cell(row=1, column=1, value=f"Производственный отчёт {period.title}")
    if max_width > 1:
        summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_width)
    header_rows: set[int] = set()
    section_rows: set[int] = set()
    current_row = 3
    for title, header, rows in sections:
        section_row, header_row, current_row = _append_report_section(summary, title, header, rows, current_row)
        section_rows.add(section_row)
        header_rows.add(header_row)
    _style_used_range(summary, header_rows=header_rows, section_rows=section_rows)

    used_names = {summary.title}
    for title, header, rows in sections:
        base = title[:31] or "Раздел"
        sheet_name = base
        index = 2
        while sheet_name in used_names:
            suffix = f" {index}"
            sheet_name = (base[:31 - len(suffix)] + suffix)[:31]
            index += 1
        used_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)
        ws.append(header)
        for row in rows or [_empty_row(len(header))]:
            ws.append(row)
        for col_index, header_name in enumerate(header, start=1):
            number_format = _number_format_for_header(str(header_name))
            for cell in ws.iter_cols(min_col=col_index, max_col=col_index, min_row=2, max_row=ws.max_row):
                for item in cell:
                    item.number_format = number_format
        _style_sheet(ws)

    filename = f"uchet_{_safe_name(period.title)}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    path = reports_dir() / filename
    wb.save(path)
    return path

def _stringify_table(header: list[str], rows: list[list[object]]) -> list[list[str]]:
    data: list[list[str]] = [[str(cell) for cell in header]]
    if rows:
        data.extend([[_display_cell(cell) for cell in row] for row in rows])
    else:
        data.append(_empty_row(len(header)))
    return data


def _pdf_cell(value: object, style: ParagraphStyle) -> Paragraph:
    return Paragraph(html.escape("" if value is None else str(value)).replace("\n", "<br/>"), style)



def _pdf_col_widths(data: list[list[str]]) -> list[float]:
    page_width = landscape(A3)[0] - 16 * mm
    if not data:
        return []
    col_count = len(data[0])
    if col_count <= 0:
        return []
    weights: list[float] = []
    for index in range(col_count):
        sample = max((len(str(row[index])) if index < len(row) else 0 for row in data[:90]), default=8)
        if index == 0:
            weights.append(max(16.0, min(34.0, float(sample))))
        else:
            weights.append(max(8.0, min(26.0, float(sample))))
    total = sum(weights) or 1.0
    return [page_width * w / total for w in weights]


def _pdf_table(data: list[list[str]], font_name: str) -> Table:
    col_count = len(data[0]) if data else 1
    font_size = 8.2 if col_count <= 7 else 7.0
    leading = font_size + 2.2
    cell_style = ParagraphStyle("CellRu", fontName=font_name, fontSize=font_size, leading=leading, wordWrap="CJK")
    header_style = ParagraphStyle("HeaderRu", parent=cell_style, fontName=font_name, fontSize=font_size, leading=leading, wordWrap="CJK")
    converted: list[list[Paragraph]] = []
    for row_index, row in enumerate(data):
        style = header_style if row_index == 0 else cell_style
        converted.append([_pdf_cell(cell, style) for cell in row])
    table = Table(converted, repeatRows=1, colWidths=_pdf_col_widths(data), splitByRow=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return table


def _pdf_section_tables(title: str, header: list[str], rows: list[list[object]], font_name: str) -> list[Table]:
    if len(header) <= 12:
        return [_pdf_table(_stringify_table(header, rows), font_name)]
    fixed_count = min(3, len(header))
    fixed_header = header[:fixed_count]
    tail_header = header[fixed_count:]
    total_header = []
    if tail_header and tail_header[-1] == "Итого":
        total_header = [tail_header[-1]]
        tail_header = tail_header[:-1]
    tables: list[Table] = []
    chunk_size = 8
    for index in range(0, len(tail_header), chunk_size):
        chunk = tail_header[index:index + chunk_size]
        new_header = [*fixed_header, *chunk, *total_header]
        new_rows: list[list[object]] = []
        for row in rows:
            fixed_values = row[:fixed_count]
            chunk_values = row[fixed_count + index:fixed_count + index + len(chunk)]
            total_values = [row[-1]] if total_header and row else []
            new_rows.append([*fixed_values, *chunk_values, *total_values])
        tables.append(_pdf_table(_stringify_table(new_header, new_rows), font_name))
    return tables or [_pdf_table(_stringify_table(header, rows), font_name)]


def create_pdf_report(chat_id: int, request_text: str = "отчёт", user_id: int | None = None) -> Path:
    period = _date_bounds_for_text(request_text)
    if period.error:
        raise ValueError(period.error)
    font_name = _register_pdf_font()
    filename = f"uchet_{_safe_name(period.title)}_{datetime.now():%Y%m%d_%H%M%S}.pdf"
    path = reports_dir() / filename
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A3), rightMargin=8*mm, leftMargin=8*mm, topMargin=8*mm, bottomMargin=8*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleRuReadable", parent=styles["Title"], fontName=font_name, fontSize=16, leading=20)
    section_style = ParagraphStyle("SectionRuReadable", parent=styles["Heading2"], fontName=font_name, fontSize=11, leading=14, spaceBefore=6, spaceAfter=4)
    small_style = ParagraphStyle("SmallRuReadable", parent=styles["Normal"], fontName=font_name, fontSize=8, leading=10)
    story = [Paragraph(f"Производственный отчёт {period.title}", title_style), Spacer(1, 6)]

    sections = report_sections(chat_id, request_text, user_id=user_id, selected=_full_export_selected())
    if not sections:
        sections = [("Отчёт", ["Раздел", "Состояние"], [["Отчёт", "Нет данных"]])]
    for section_index, (title, header, rows) in enumerate(sections):
        if section_index and title in {"По датам", "Журнал"}:
            story.append(PageBreak())
        story.append(Paragraph(title, section_style))
        if title == "По датам" and len(header) > 12:
            story.append(Paragraph("Широкая таблица разделена на удобные части.", small_style))
            story.append(Spacer(1, 3))
        tables = _pdf_section_tables(title, header, rows or [_empty_row(len(header))], font_name)
        for table_index, table in enumerate(tables):
            if table_index:
                story.append(Spacer(1, 6))
            story.append(table)
        story.append(Spacer(1, 8))
    doc.build(story)
    return path



# --- Разделы отчёта для Excel и PDF ---


def _period_for_export(request_text: str) -> PeriodFilter:
    period = _date_bounds_for_text(request_text)
    if period.error:
        raise ValueError(period.error)
    return period


def _inventory_table(chat_id: int) -> list[list[object]]:
    return [[
        ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or ""),
        row.get("entity_name") or "",
        row.get("area_name") or "Общий склад",
        float(row.get("quantity") or 0),
        row.get("unit") or "",
    ] for row in inventory_rows(chat_id)]


def _period_totals_table(chat_id: int, period: PeriodFilter) -> list[list[object]]:
    return [[
        OPERATION_LABELS.get(row.get("operation_type") or "", row.get("operation_type") or ""),
        ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or ""),
        row.get("entity_name") or "",
        row.get("area_name") or "",
        float(row.get("total_quantity") or 0),
        row.get("unit") or "",
        int(row.get("count_rows") or 0),
    ] for row in operation_rows(chat_id, period)]


DEFAULT_ASSEMBLY_TARGETS = [10000, 50000, 100000]


_TARGET_NUMBER_RE = re.compile(r"(?<!\d)(\d{1,3}(?:[ \u00A0]\d{3})+|\d{3,12})(?!\d)")
_TARGET_MARKED_NUMBER_RE = re.compile(
    r"(?:план|цель|цели|целью|сборки|сборку|собрать|сбора|до|для|на)\s+"
    r"(\d{1,3}(?:[ \u00A0]\d{3})+|\d{1,12})(?!\d)",
    re.IGNORECASE,
)


def _normalize_target_number(raw: str) -> int | None:
    clean = re.sub(r"[ \u00A0]+", "", raw or "")
    if not clean.isdigit():
        return None
    value = int(clean)
    return value if value > 0 else None


def _target_quantities_from_text(text: str, include_defaults: bool = True) -> list[int]:
    cleaned = _DATE_PATTERN.sub(" ", text or "")
    found: list[int] = []
    for pattern in (_TARGET_NUMBER_RE, _TARGET_MARKED_NUMBER_RE):
        for match in pattern.finditer(cleaned):
            value = _normalize_target_number(match.group(1))
            if value and value not in found:
                found.append(value)
    if found:
        return found
    return list(DEFAULT_ASSEMBLY_TARGETS) if include_defaults else []


def _strip_target_numbers(text: str) -> str:
    cleaned = _DATE_PATTERN.sub(" ", text or "")
    cleaned = _TARGET_MARKED_NUMBER_RE.sub(" ", cleaned)
    cleaned = _TARGET_NUMBER_RE.sub(" ", cleaned)
    return cleaned


def _custom_product_targets_from_text(chat_id: int, request_text: str) -> dict[int, list[int]]:
    """Возвращает цели, заданные в запросе для конкретных изделий.

    Пример: "план сборки: Изделие 1 50000; Изделие 2 120000".
    """
    from .matcher import confident_match

    chunks = [chunk.strip() for chunk in re.split(r"[;,\n]+", request_text or "") if chunk.strip()]
    result: dict[int, list[int]] = {}
    for chunk in chunks:
        targets = _target_quantities_from_text(chunk, include_defaults=False)
        if not targets:
            continue
        product_text = _strip_target_numbers(chunk)
        match, _ = confident_match(chat_id, product_text.strip() or chunk, allowed_types={"product"})
        if not match:
            continue
        bucket = result.setdefault(int(match.target_id), [])
        for target in targets:
            if target not in bucket:
                bucket.append(target)
    return result

def _component_stock_rows(chat_id: int) -> list[list[object]]:
    rows = db.fetchall(
        """
        SELECT i.entity_type,e.name AS entity_name,i.unit,COALESCE(SUM(i.quantity),0) AS quantity
        FROM inventory i
        JOIN entities e ON e.id=i.entity_id
        WHERE i.chat_id=? AND i.entity_type IN ('component','stock_item') AND e.is_archived=0
        GROUP BY i.entity_type,i.entity_id,i.unit
        ORDER BY i.entity_type,e.name,i.unit
        """,
        (chat_id,),
    )
    result: list[list[object]] = []
    for raw_row in rows:
        row = dict(raw_row)
        result.append([
            ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or ""),
            row.get("entity_name") or "",
            float(row.get("quantity") or 0),
            row.get("unit") or "",
        ])
    return result


def _component_daily_table(chat_id: int, period: PeriodFilter) -> tuple[list[str], list[list[object]]]:
    bucket_type, labels = _bucket_labels_for_period(period)
    bucket_sql = "strftime('%d.%m.%Y', o.created_at)" if bucket_type == "day" else "strftime('%m.%Y', o.created_at)"
    rows = db.fetchall(
        f"""
        SELECT o.entity_type,e.name AS entity_name,o.unit,{bucket_sql} AS bucket,
               SUM(o.quantity) AS total_quantity
        FROM operations o
        JOIN entities e ON e.id=o.entity_id
        WHERE o.chat_id=? AND {period.where_sql}
          AND o.operation_type='production'
          AND o.entity_type IN ('component','stock_item')
          AND e.is_archived=0
        GROUP BY o.entity_type,o.entity_id,o.unit,bucket
        ORDER BY o.entity_type,e.name,bucket
        """,
        (chat_id, *period.params),
    )
    grouped: dict[tuple[str, str, str], dict[str, object]] = {}
    for raw_row in rows:
        row = dict(raw_row)
        key = (str(row.get("entity_type") or ""), str(row.get("entity_name") or ""), str(row.get("unit") or ""))
        item = grouped.setdefault(key, {"entity_type": key[0], "entity_name": key[1], "unit": key[2], "values": {label: 0.0 for label in labels}})
        bucket = str(row.get("bucket") or "")
        if bucket in item["values"]:
            item["values"][bucket] = float(row.get("total_quantity") or 0)
    table: list[list[object]] = []
    for item in grouped.values():
        values = [float(item["values"].get(label) or 0) for label in labels]
        table.append([
            ENTITY_LABELS.get(str(item.get("entity_type") or ""), str(item.get("entity_type") or "")),
            item.get("entity_name") or "",
            item.get("unit") or "",
            *values,
            sum(values),
        ])
    return labels, table


def _assembly_shipment_daily_table(chat_id: int, period: PeriodFilter) -> tuple[list[str], list[list[object]]]:
    bucket_type, labels = _bucket_labels_for_period(period)
    bucket_sql = "strftime('%d.%m.%Y', o.created_at)" if bucket_type == "day" else "strftime('%m.%Y', o.created_at)"
    rows = db.fetchall(
        f"""
        SELECT o.operation_type,e.name AS entity_name,o.unit,{bucket_sql} AS bucket,
               SUM(o.quantity) AS total_quantity
        FROM operations o
        JOIN entities e ON e.id=o.entity_id
        WHERE o.chat_id=? AND {period.where_sql}
          AND o.operation_type IN ('assembly','shipment')
          AND o.entity_type='product'
          AND e.is_archived=0
        GROUP BY o.operation_type,o.entity_id,o.unit,bucket
        ORDER BY o.operation_type,e.name,bucket
        """,
        (chat_id, *period.params),
    )
    grouped: dict[tuple[str, str, str], dict[str, object]] = {}
    for raw_row in rows:
        row = dict(raw_row)
        key = (str(row.get("operation_type") or ""), str(row.get("entity_name") or ""), str(row.get("unit") or ""))
        item = grouped.setdefault(key, {"operation_type": key[0], "entity_name": key[1], "unit": key[2], "values": {label: 0.0 for label in labels}})
        bucket = str(row.get("bucket") or "")
        if bucket in item["values"]:
            item["values"][bucket] = float(row.get("total_quantity") or 0)
    table: list[list[object]] = []
    for item in grouped.values():
        values = [float(item["values"].get(label) or 0) for label in labels]
        table.append([
            OPERATION_LABELS.get(str(item.get("operation_type") or ""), str(item.get("operation_type") or "")),
            item.get("entity_name") or "",
            item.get("unit") or "",
            *values,
            sum(values),
        ])
    return labels, table


def _product_target_rows(
    chat_id: int,
    targets: list[int],
    product_ids: set[int] | None = None,
    per_product_targets: dict[int, list[int]] | None = None,
) -> list[list[object]]:
    from . import repository as repo

    rows: list[list[object]] = []
    for product_pack in repo.all_products_with_components(chat_id):
        product = product_pack["product"]
        product_id = int(product.id)
        if product_ids is not None and product_id not in product_ids:
            continue
        product_targets = list(per_product_targets.get(product_id, []) if per_product_targets else targets)
        if not product_targets:
            continue
        components = product_pack["components"]
        if not components:
            for target in product_targets:
                rows.append([product.name, target, "состав не задан", "", "", "", "", "", ""])
            continue
        possible_values: list[float] = []
        component_info: list[dict[str, object]] = []
        for comp in components:
            stock_qty = _component_stock(chat_id, int(comp["component_id"]))
            need = float(comp["quantity"] or 0)
            possible_values.append(stock_qty // need if need > 0 else 0)
            component_info.append({
                "name": str(comp.get("name") or ""),
                "stock": stock_qty,
                "need": need,
                "unit": str(comp.get("default_unit") or "шт"),
            })
        possible = int(min(possible_values)) if possible_values else 0
        for target in product_targets:
            for comp in component_info:
                need_total = float(comp["need"]) * target
                missing = max(0.0, need_total - float(comp["stock"]))
                rows.append([
                    product.name,
                    target,
                    possible,
                    comp["name"],
                    comp["stock"],
                    comp["need"],
                    need_total,
                    missing,
                    comp["unit"],
                ])
    return rows


def _product_target_rows_for_request(chat_id: int, request_text: str, include_defaults: bool = True) -> list[list[object]]:
    from .matcher import confident_match

    custom = _custom_product_targets_from_text(chat_id, request_text)
    if custom:
        return _product_target_rows(chat_id, [], product_ids=set(custom), per_product_targets=custom)

    targets = _target_quantities_from_text(request_text, include_defaults=include_defaults)
    product_text = _strip_target_numbers(request_text)
    match, _ = confident_match(chat_id, product_text.strip() or request_text, allowed_types={"product"})
    if match:
        return _product_target_rows(chat_id, targets, product_ids={int(match.target_id)})
    return _product_target_rows(chat_id, targets)

def _assembly_shipping_summary_table(chat_id: int, period: PeriodFilter) -> list[list[object]]:
    rows = db.fetchall(
        f"""
        SELECT e.name AS entity_name,o.unit,
               SUM(CASE WHEN o.operation_type='assembly' THEN o.quantity ELSE 0 END) AS assembled,
               SUM(CASE WHEN o.operation_type='shipment' THEN o.quantity ELSE 0 END) AS shipped
        FROM operations o
        JOIN entities e ON e.id=o.entity_id
        WHERE o.chat_id=? AND {period.where_sql}
          AND o.operation_type IN ('assembly','shipment')
          AND o.entity_type='product'
          AND e.is_archived=0
        GROUP BY o.entity_id,o.unit
        ORDER BY e.name,o.unit
        """,
        (chat_id, *period.params),
    )
    result: list[list[object]] = []
    for raw_row in rows:
        row = dict(raw_row)
        result.append([
            row.get("entity_name") or "",
            float(row.get("assembled") or 0),
            float(row.get("shipped") or 0),
            row.get("unit") or "",
        ])
    return result


def _capacity_table(chat_id: int) -> list[list[object]]:
    return [[
        row.get("product_name") or "",
        row.get("possible") if row.get("possible") != "" else "",
        row.get("component_name") or "",
        row.get("stock") if row.get("stock") != "" else "",
        row.get("need") if row.get("need") != "" else "",
        row.get("missing_next") if row.get("missing_next") != "" else "",
        row.get("unit") or "",
    ] for row in product_capacity_rows(chat_id)]


def _journal_table(chat_id: int, period: PeriodFilter, limit: int = 5000) -> list[list[object]]:
    return [[
        row.get("created_at") or "",
        OPERATION_LABELS.get(row.get("operation_type") or "", row.get("operation_type") or ""),
        ENTITY_LABELS.get(row.get("entity_type") or "", row.get("entity_type") or ""),
        row.get("entity_name") or "",
        row.get("area_name") or "",
        float(row.get("quantity") or 0),
        row.get("unit") or "",
        row.get("user_id") or "",
        row.get("group_chat_id") or "",
    ] for row in raw_operation_rows(chat_id, period, limit=limit)]


def _daily_matrix_table(chat_id: int, period: PeriodFilter) -> tuple[list[str], list[list[object]]]:
    labels, matrix = movement_matrix_rows(chat_id, period)
    rows: list[list[object]] = []
    for row in matrix:
        values = [float(row["values"].get(label) or 0) for label in labels]
        rows.append([
            OPERATION_LABELS.get(row.get("operation_type") or "", row.get("operation_type") or ""),
            row.get("entity_name") or ENTITY_LABELS.get(row.get("entity_type") or "", ""),
            row.get("unit") or "",
            *values,
            sum(values),
        ])
    return labels, rows


def report_sections(
    chat_id: int,
    request_text: str = "отчёт",
    user_id: int | None = None,
    selected: dict[str, bool] | None = None,
) -> list[tuple[str, list[str], list[list[object]]]]:
    from . import repository as repo

    period = _period_for_export(request_text)
    prefs = selected if selected is not None else repo.get_export_preferences(chat_id, user_id)
    sections: list[tuple[str, list[str], list[list[object]]]] = []
    if prefs.get("inventory"):
        sections.append(("Склад", ["Тип", "Название", "Участок", "Количество", "Ед."], _inventory_table(chat_id)))
        sections.append(("Остатки комплектующих", ["Тип", "Название", "Остаток", "Ед."], _component_stock_rows(chat_id)))
    if prefs.get("period_totals"):
        sections.append(("Итоги за период", ["Операция", "Тип", "Название", "Участок", "Количество", "Ед.", "Строк"], _period_totals_table(chat_id, period)))
    if prefs.get("daily_matrix"):
        labels, rows = _daily_matrix_table(chat_id, period)
        sections.append(("По датам", ["Операция", "Название", "Ед.", *labels, "Итого"], rows))
        comp_labels, comp_rows = _component_daily_table(chat_id, period)
        sections.append(("Комплектующие по датам", ["Тип", "Название", "Ед.", *comp_labels, "Итого"], comp_rows))
        move_labels, move_rows = _assembly_shipment_daily_table(chat_id, period)
        sections.append(("Сборка и отправка по датам", ["Операция", "Изделие", "Ед.", *move_labels, "Итого"], move_rows))
    if prefs.get("capacity"):
        sections.append(("Расчёт сборки", ["Изделие", "Можно собрать", "Комплектующая", "Есть", "Нужно на 1", "Не хватает для ещё 1", "Ед."], _capacity_table(chat_id)))
        sections.append(("План сборки", ["Изделие", "Цель", "Можно собрать", "Комплектующая", "Есть", "Нужно на 1", "Нужно на цель", "Не хватает", "Ед."], _product_target_rows_for_request(chat_id, request_text, include_defaults=True)))
        sections.append(("Собрано и отправлено", ["Изделие", "Собрано", "Отправлено/продано", "Ед."], _assembly_shipping_summary_table(chat_id, period)))
    if prefs.get("journal"):
        sections.append(("Журнал", ["Дата", "Операция", "Тип", "Название", "Участок", "Количество", "Ед.", "Работник", "Группа"], _journal_table(chat_id, period)))
    return sections


def _empty_row(width: int, text: str = "Нет данных") -> list[object]:
    return [text, *["" for _ in range(max(0, width - 1))]]
